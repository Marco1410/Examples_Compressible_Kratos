import os
import time
import subprocess
import openpyxl
import numpy as np
import matplotlib
import matplotlib.pyplot as plt
matplotlib.use('Agg')
from rbf_hhrom_output import BuildRBFoutput # type: ignore
from rbf_hhrom_output import BuildHHROMoutput # type: ignore
from ClearAll import Clear # type: ignore
from scipy.stats import qmc  
import KratosMultiphysics
import KratosMultiphysics.kratos_utilities
import KratosMultiphysics.CompressiblePotentialFlowApplication as CPFApp
from KratosMultiphysics.RomApplication.rom_manager import RomManager
from KratosMultiphysics.RomApplication.calculate_rom_basis_output_process import CalculateRomBasisOutputProcess


# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
#
# get multiple parameters by Halton or LatinHypercube methods
#
def get_multiple_parameters(number_train_values=0, number_test_values=0, angle=[], mach=[], method='Halton'):
    if method == 'Halton':
        sampler = qmc.Halton(d=2)
    elif method == 'LatinHypercube':
        sampler = qmc.LatinHypercube(d=2)
    mu_train = []; mu_test = []; mu_train_not_scaled = []; mu_test_not_scaled = []
    if number_train_values > 0:
        sample = sampler.random(number_train_values)
        values = qmc.scale(sample, [angle[0],mach[0]], [angle[1],mach[1]])
        for i in range(number_train_values):
            #Angle of attack , Mach infinit
            mu_train.append([values[i,0], values[i,1]])
            mu_train_not_scaled.append([sample[i,0], sample[i,1]])
        np.save(f'mu_train', mu_train)
        np.save(f'mu_train_not_scaled', mu_train_not_scaled)
    if number_test_values > 0:
        sample = sampler.random(number_test_values)
        values = qmc.scale(sample, [angle[0],mach[0]], [angle[1],mach[1]])
        for i in range(number_test_values):
            #Angle of attack , Mach infinit
            mu_test.append([values[i,0], values[i,1]])
            mu_test_not_scaled.append([sample[i,0], sample[i,1]])
        np.save(f'mu_test', mu_test)
        np.save(f'mu_test_not_scaled', mu_test_not_scaled)
    
    plot_mu_values(mu_train, mu_test, 'MuValues')
    plot_mu_values(mu_train_not_scaled, mu_test_not_scaled, 'MuValuesNotScaled')
    
    return mu_train, mu_test, mu_train_not_scaled, mu_test_not_scaled
    
# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
#
# wake angles Salome
#
def SaveAngles(mu_list):
    fout = open('wake_angles.dat','w')
    for mu in mu_list:
        fout.write("%s\n" %(mu[0]))
        if len(mu_list) == 1:
            fout.write("0.0 \n")
    fout.close()

def LaunchSalome():
    salome_cmd = "salome -t python"
    salome_script_name = "wake_salome.py"
    salome_exe = " ".join([salome_cmd, salome_script_name])
    subprocess.Popen(["/bin/bash", "-i", "-c", salome_exe])

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
#
# plot parameters
#
def plot_mu_values(mu_train, mu_test, name):
    if len(mu_train) > 0: plt.plot(np.array(mu_train)[:,1], np.array(mu_train)[:,0], 'bs', label="Train Values")
    if len(mu_test ) > 0: plt.plot(np.array( mu_test)[:,1], np.array( mu_test)[:,0], 'ro', label="Test Values")
    plt.title('Mu Values')
    plt.ylabel('Alpha')
    plt.xlabel('Mach')
    plt.grid(True)
    plt.legend(bbox_to_anchor=(.85, 1.03, 1., .102), loc='upper left', borderaxespad=0.)
    plt.savefig(f"{name}.png")
    plt.close('all')

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
#
# load parameters
#
def load_mu_parameters():
    if os.path.exists("mu_train.npy") and os.path.exists("mu_test.npy"):
        mu_train = np.load('mu_train.npy')
        mu_train_not_scaled = np.load('mu_train_not_scaled.npy')
        mu_test = np.load('mu_test.npy')
        mu_test_not_scaled = np.load('mu_test_not_scaled.npy')
        mu_train =  [mu.tolist() for mu in mu_train]
        mu_train_not_scaled =  [mu.tolist() for mu in mu_train_not_scaled]
        mu_test =  [mu.tolist() for mu in mu_test]
        mu_test_not_scaled =  [mu.tolist() for mu in mu_test_not_scaled]
    elif os.path.exists("mu_train.npy"):
        mu_train = np.load('mu_train.npy')
        mu_train_not_scaled = np.load('mu_train_not_scaled.npy')
        mu_train =  [mu.tolist() for mu in mu_train]
        mu_train_not_scaled =  [mu.tolist() for mu in mu_train_not_scaled]
        mu_test = []
        mu_test_not_scaled = []
    elif os.path.exists("mu_test.npy"):
        mu_test = np.load('mu_test.npy')
        mu_test_not_scaled = np.load('mu_test_not_scaled.npy')
        mu_test =  [mu.tolist() for mu in mu_test]
        mu_test_not_scaled =  [mu.tolist() for mu in mu_test_not_scaled]
        mu_train = []
        mu_train_not_scaled = []
    return mu_train, mu_test, mu_train_not_scaled, mu_test_not_scaled

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
#
# RBF prediction and error
#
def RBF_prediction( mu_train                 = [None],
                    mu_test                  = [None],
                    mu_train_not_scaled      = [None],
                    mu_test_not_scaled       = [None]):
    from ezyrb import ReducedOrderModel as ROM
    from ezyrb import RBF, POD, Database

    #### CLUSTERING DATA
    #################################################################################################
    parameters = mu_train_not_scaled

    snapshots = []
    for mu in mu_train:
        file = f'{mu[0]}, {mu[1]}.npy'
        snapshots.append(np.load(f'FOM_Snapshots/{file}'))
    snapshots = np.block(snapshots)

    #### RBF TRAINING
    #################################################################################################
    db = Database(parameters, snapshots.T)
    pod = POD()
    rbf = RBF()
    rom = ROM(db, pod, rbf).fit()

    if len(mu_test) > 0:
        #### PREDICTION OF TEST
        #################################################################################################
        interpolated_solutions_list = [rom.predict([element]).snapshots_matrix for element in mu_test_not_scaled]

        for i, solution in enumerate(interpolated_solutions_list):
            np.save(f"RBF_Snapshots/{mu_test[i][0]}, {mu_test[i][1]}.npy", solution)

def PEBL_error_estimation(mu_train, mu_test):
    if len(mu_train) > 0:
        approximation_error = 0.0
        FOM_model = []; RBF_model = []
        for mu in mu_train:
            FOM_model.append(np.load(f'FOM_Snapshots/{mu[0]}, {mu[1]}.npy'))
            RBF_model.append(np.load(f"RBF_Snapshots/{mu[0]}, {mu[1]}.npy").T)
        FOM_model = np.block(FOM_model)
        RBF_model = np.block(RBF_model)
        training_approximation_error = np.linalg.norm(FOM_model - RBF_model)/np.linalg.norm(FOM_model)
        print(f'RBF training approximation error: {training_approximation_error:.2E}')

    if len(mu_test)>0 and os.path.exists(f'RBF_Snapshots/{mu_test[0][0]}, {mu_test[0][1]}.npy'):
        approximation_error = 0.0
        FOM_model = []; RBF_model_interpolation = []
        for mu in mu_test:
            FOM_model.append(np.load(f'FOM_Snapshots/{mu[0]}, {mu[1]}.npy'))
            RBF_model_interpolation.append(np.load(f"RBF_Snapshots/{mu[0]}, {mu[1]}.npy").T)
        FOM_model = np.block(FOM_model)
        RBF_model_interpolation = np.block(RBF_model_interpolation)
        approximation_error = np.linalg.norm(FOM_model - RBF_model_interpolation)/np.linalg.norm(FOM_model)
        print(f'RBF interpolation approximation error: {approximation_error:.2E}')

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #


def CustomizeSimulation(cls, global_model, parameters):

    class CustomSimulation(cls):

        def __init__(self, model,project_parameters):
            super().__init__(model,project_parameters)
        
        def Initialize(self):
            super().Initialize()

        def Run(self):

            angle = parameters["processes"]["boundary_conditions_process_list"][0]["Parameters"]["angle_of_attack"].GetDouble()
            mach  = parameters["processes"]["boundary_conditions_process_list"][0]["Parameters"]["mach_infinity"].GetDouble()

            case_name = f'{angle}, {mach}'

            info_steps_list = []
            error = 0
            modes = 0

            if self._GetSimulationName() == "Analysis": # FOM

                start_time = time.time()
                self.Initialize()
                self.RunSolutionLoop()
                self.Finalize()

                exe_time = time.time() - start_time

                if parameters["output_processes"].Has("gid_output"):

                    simulation_name = parameters["output_processes"]["gid_output"][0]["Parameters"]["output_name"].GetString().removeprefix('Results/')

                    for process in self._GetListOfOutputProcesses():
                            if isinstance(process, CalculateRomBasisOutputProcess):
                                BasisOutputProcess = process

                    if 'Fit' in simulation_name:
                        case_type = 'train_fom'
                    elif 'Test' in simulation_name:
                        case_type = 'test_fom' 
                        modes = np.load('rom_data/RightBasisMatrix.npy').shape[1]
                    fom = BasisOutputProcess._GetSnapshotsMatrix()
                    np.save(f'FOM_Snapshots/{case_name}',fom)

                    modelpart = self.model["MainModelPart"]
                    fout = open("trailing_edge_element_id.txt",'a')
                    for elem in modelpart.Elements:
                        if elem.GetValue(CPFApp.TRAILING_EDGE):
                            fout.write("%s\n" %(elem.Id))
                        if elem.GetValue(CPFApp.KUTTA):
                            fout.write("%s\n" %(elem.Id))
                    fout.close()

                    fout = open("upwind_elements_list.txt",'a')
                    for elem in modelpart.Elements:
                        if elem.GetValue(CPFApp.ID_UPWIND_ELEMENT) != 0.0:
                            fout.write("%s\n" %(elem.Id))
                            fout.write("%s\n" %(elem.GetValue(CPFApp.ID_UPWIND_ELEMENT)))
                    fout.close()
                
                    info_steps_list.append([case_type,
                                            angle,
                                            mach, 
                                            self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.NL_ITERATION_NUMBER],
                                            self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.RESIDUAL_NORM],
                                            error,
                                            modes,
                                            round(exe_time, 2),
                                            self.model["MainModelPart"].NumberOfNodes(),
                                            self.model["MainModelPart"].ProcessInfo[CPFApp.LIFT_COEFFICIENT],
                                            self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.DRAG_COEFFICIENT]])
                    
                    if os.path.exists(f'FOM_data.xlsx'):
                        wb = openpyxl.load_workbook(f'FOM_data.xlsx')
                        hoja = wb.active
                        for item in info_steps_list:
                            hoja.append(item)
                        wb.save(f'FOM_data.xlsx')
                    else:
                        wb = openpyxl.Workbook()
                        hoja = wb.active
                        hoja.append(('Case name', 'Angle [º]', 'Mach', 'NL iterations', 'Residual norm', 'Approximation error [%]', 'Modes', 'Time [sec]', 'Nodes', 'Cl', 'Cd'))
                        for item in info_steps_list:
                            hoja.append(item)
                        wb.save(f'FOM_data.xlsx')

            elif self._GetSimulationName() == "::[ROM Simulation]:: ": # ROM

                start_time = time.time()
                self.Initialize()
                self.RunSolutionLoop()
                self.Finalize()                
                exe_time = time.time() - start_time

                if parameters["output_processes"].Has("gid_output"):

                    simulation_name = parameters["output_processes"]["gid_output"][0]["Parameters"]["output_name"].GetString().removeprefix('Results/')

                    for process in self._GetListOfOutputProcesses():
                            if isinstance(process, CalculateRomBasisOutputProcess):
                                BasisOutputProcess = process

                    if 'HROM' in simulation_name:                      # HROM
                        if 'Fit' in simulation_name:
                            case_type = 'train_'
                        elif 'Test' in simulation_name:
                            case_type = 'test_'
                            
                        hrom = BasisOutputProcess._GetSnapshotsMatrix()
                        fom = np.load(f'FOM_Snapshots/{case_name}.npy')

                        if (len(fom) != len(hrom)):
                            q_matrix = []
                            main_model_part = self.model["MainModelPart"]
                            q_matrix.append(np.array(main_model_part.GetValue(KratosMultiphysics.RomApplication.ROM_SOLUTION_INCREMENT)).reshape(-1,1))
                            q_matrix = np.block(q_matrix)
                            phi = np.load(f'rom_data/RightBasisMatrix.npy')
                            if (q_matrix.shape[0] == 1): q_matrix = q_matrix.T
                            hrom = phi @ q_matrix
                            np.save(f'HHROM_Snapshots/{case_name}', hrom)
                            case_type = case_type + 'hhrom'
                        else:
                            np.save(f'HROM_Snapshots/{case_name}', hrom)
                            case_type = case_type + 'hrom'

                        error = np.linalg.norm(fom-hrom)/np.linalg.norm(fom)
                        modes = np.load('rom_data/RightBasisMatrix.npy').shape[1]
                    
                        info_steps_list.append([case_type,
                                                angle,
                                                mach, 
                                                self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.NL_ITERATION_NUMBER],
                                                self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.RESIDUAL_NORM],
                                                error,
                                                modes,
                                                round(exe_time, 2),
                                                self.model["MainModelPart"].NumberOfNodes(),
                                                self.model["MainModelPart"].ProcessInfo[CPFApp.LIFT_COEFFICIENT],
                                                self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.DRAG_COEFFICIENT]])
                        
                        if os.path.exists(f'case_data.xlsx'):
                            wb = openpyxl.load_workbook(f'case_data.xlsx')
                            hoja = wb.active
                            for item in info_steps_list:
                                hoja.append(item)
                            wb.save(f'case_data.xlsx')
                        else:
                            wb = openpyxl.Workbook()
                            hoja = wb.active
                            hoja.append(('Case name', 'Angle [º]', 'Mach', 'NL iterations', 'Residual norm', 'Approximation error [%]', 'Modes', 'Time [sec]', 'Nodes', 'Cl', 'Cd'))
                            for item in info_steps_list:
                                hoja.append(item)
                            wb.save(f'case_data.xlsx')

                    elif 'ROM' in simulation_name:                    # ROM
                        if 'Fit' in simulation_name:
                            case_type = 'train_rom'
                        elif 'Test' in simulation_name:
                            case_type = 'test_rom' 
                        modes = np.load('rom_data/RightBasisMatrix.npy').shape[1]

                        rom = BasisOutputProcess._GetSnapshotsMatrix()
                        np.save(f'ROM_Snapshots/{case_name}',rom)
                        fom = np.load(f'FOM_Snapshots/{case_name}.npy')
                        error = np.linalg.norm(fom-rom)/np.linalg.norm(fom)
                    
                        info_steps_list.append([case_type,
                                                angle,
                                                mach, 
                                                self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.NL_ITERATION_NUMBER],
                                                self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.RESIDUAL_NORM],
                                                error,
                                                modes,
                                                round(exe_time, 2),
                                                self.model["MainModelPart"].NumberOfNodes(),
                                                self.model["MainModelPart"].ProcessInfo[CPFApp.LIFT_COEFFICIENT],
                                                self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.DRAG_COEFFICIENT]])
                        
                        if os.path.exists(f'case_data.xlsx'):
                            wb = openpyxl.load_workbook(f'case_data.xlsx')
                            hoja = wb.active
                            for item in info_steps_list:
                                hoja.append(item)
                            wb.save(f'case_data.xlsx')
                        else:
                            wb = openpyxl.Workbook()
                            hoja = wb.active
                            hoja.append(('Case name','Angle [º]', 'Mach', 'NL iterations', 'Residual norm', 'Approximation error [%]', 'Modes', 'Time [sec]', 'Nodes', 'Cl', 'Cd'))
                            for item in info_steps_list:
                                hoja.append(item)
                            wb.save(f'case_data.xlsx')

    return CustomSimulation(global_model, parameters)


# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

def UpdateProjectParameters(parameters, mu=None):
    angle_of_attack = mu[0]
    mach_infinity   = mu[1]
    wake_normal     = [-np.sin(angle_of_attack*np.pi/180),0.0,np.cos(angle_of_attack*np.pi/180)]
    parameters["processes"]["boundary_conditions_process_list"][0]["Parameters"]["angle_of_attack"].SetDouble(np.double(angle_of_attack))
    parameters["processes"]["boundary_conditions_process_list"][0]["Parameters"]["mach_infinity"].SetDouble(np.double(mach_infinity))
    parameters["processes"]["boundary_conditions_process_list"][1]["Parameters"]["wake_stl_file_name"].SetString(f'SalomeFiles/Wake_{angle_of_attack}.stl')
    parameters["processes"]["boundary_conditions_process_list"][1]["Parameters"]["wake_process_cpp_parameters"]["wake_normal"].SetVector(wake_normal)

    return parameters

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

def UpdateMaterialParametersFile(material_parametrs_file_name, mu):
    pass

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #



def GetRomManagerParameters():
    general_rom_manager_parameters = KratosMultiphysics.Parameters("""{
            "rom_stages_to_train" : ["ROM","HROM"],            // ["ROM","HROM"]
            "rom_stages_to_test"  : ["ROM","HROM"],            // ["ROM","HROM"]
            "paralellism" : null,                       // null, TODO: add "compss"
            "projection_strategy": "galerkin",          // "lspg", "galerkin", "petrov_galerkin"
            "assembling_strategy": "global",            // "global", "elemental"
            "save_gid_output": true,                    // false, true #if true, it must exits previously in the ProjectParameters.json
            "save_vtk_output": true,                   // false, true #if true, it must exits previously in the ProjectParameters.json
            "output_name": "id",                        // "id" , "mu"
            "ROM":{
                "svd_truncation_tolerance": 0,
                "model_part_name": "MainModelPart",                            // This changes depending on the simulation: Structure, FluidModelPart, ThermalPart #TODO: Idenfity it automatically
                "nodal_unknowns": ["VELOCITY_POTENTIAL", "AUXILIARY_VELOCITY_POTENTIAL"],     // Main unknowns. Snapshots are taken from these
                "rom_basis_output_format": "numpy",
                "rom_basis_output_name": "RomParameters",
                "rom_basis_output_folder": "rom_data",
                "snapshots_control_type": "step",                          // "step", "time"
                "snapshots_interval": 1,
                "galerkin_rom_bns_settings": {
                    "monotonicity_preserving": false
                },
                "lspg_rom_bns_settings": {
                    "train_petrov_galerkin": false,
                    "basis_strategy": "reactions",                        // 'residuals', 'jacobian', 'reactions'
                    "include_phi": false,
                    "svd_truncation_tolerance": 1e-9,
                    "solving_technique": "normal_equations",              // 'normal_equations', 'qr_decomposition'
                    "monotonicity_preserving": false
                },
                "petrov_galerkin_rom_bns_settings": {
                    "monotonicity_preserving": false
                }
            },
            "HROM":{
                "element_selection_type": "empirical_cubature",
                "element_selection_svd_truncation_tolerance": 0,
                "create_hrom_visualization_model_part" : true,
                "echo_level" : 1,                                       
                "hrom_format": "numpy",
                "include_conditions_model_parts_list": [],
                "include_elements_model_parts_list": ["MainModelPart.trailing_edge_element"],
                "initial_candidate_elements_model_part_list" : [],
                "initial_candidate_conditions_model_part_list" : [],
                "include_nodal_neighbouring_elements_model_parts_list":[],
                "include_minimum_condition": false,
                "include_condition_parents": true,
                "svd_type": "numpy_svd"
            }
        }""")

    return general_rom_manager_parameters


# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #


if __name__ == "__main__":

    folder_names = ["FOM_Snapshots"  , 
                    "ROM_Snapshots"  , 
                    "HROM_Snapshots" ,
                    "HHROM_Snapshots", 
                    "RBF_Snapshots"  ]
    for name in folder_names:
        if not os.path.exists(name):
            os.mkdir(name)

    ###############################
    # PARAMETERS SETTINGS
    update_parameters    = True
    use_hhrom_model_part = False
    number_of_mu_train   = 3
    number_of_mu_test    = 1
    mach_range           = [ 0.80, 0.86]
    angle_range          = [ 2.00, 2.40]
    ###############################

    if update_parameters:
        mu_train, mu_test, mu_train_not_scaled, mu_test_not_scaled = get_multiple_parameters(number_train_values = number_of_mu_train,
                                                                                            number_test_values  = number_of_mu_test , 
                                                                                            angle               = angle_range       , 
                                                                                            mach                = mach_range        , 
                                                                                            method              = 'Halton'          )
        KratosMultiphysics.kratos_utilities.DeleteFileIfExisting('upwind_elements_list.txt')
        KratosMultiphysics.kratos_utilities.DeleteFileIfExisting('trailing_edge_element_id.txt')
        SaveAngles(mu_train+mu_test)
        LaunchSalome()
        time.sleep(30)
    else:
        mu_train, mu_test, mu_train_not_scaled, mu_test_not_scaled = load_mu_parameters()

    general_rom_manager_parameters = GetRomManagerParameters()
    project_parameters_name = "ProjectParameters.json"

    rom_manager = RomManager(project_parameters_name,general_rom_manager_parameters,
                             CustomizeSimulation,UpdateProjectParameters,UpdateMaterialParametersFile,
                             relaunch_FOM=False, relaunch_ROM=False, relaunch_HROM=True)

    rom_manager.Fit(mu_train, use_hhrom_model_part = use_hhrom_model_part)
    rom_manager.Test(mu_test, use_hhrom_model_part = use_hhrom_model_part)

    if number_of_mu_train >= 3:
        RBF_prediction(mu_train = mu_train, mu_train_not_scaled = mu_train_not_scaled, 
                    mu_test = mu_train + mu_test, mu_test_not_scaled  = mu_train_not_scaled + mu_test_not_scaled)

    BuildRBFoutput(mu_train + mu_test)
    BuildHHROMoutput(mu_train + mu_test)

    rom_manager.PrintErrors(show_q_errors = True)

    if number_of_mu_train >= 3:
        print('::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::')
        PEBL_error_estimation(mu_train, mu_test)
        print('::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::')

