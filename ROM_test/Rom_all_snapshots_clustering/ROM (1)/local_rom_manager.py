import KratosMultiphysics
from KratosMultiphysics.RomApplication.rom_testing_utilities import SetUpSimulationInstance
from KratosMultiphysics.RomApplication.calculate_rom_basis_output_process import CalculateRomBasisOutputProcess
from KratosMultiphysics.RomApplication.randomized_singular_value_decomposition import RandomizedSingularValueDecomposition
import numpy as np
import os
import importlib
import json
import openpyxl
from pathlib import Path


class LocalRomManager(object):

    def __init__(self,project_parameters_name, general_rom_manager_parameters, CustomizeSimulation, UpdateProjectParameters):
        self.project_parameters_name = project_parameters_name
        self.general_rom_manager_parameters = general_rom_manager_parameters
        self.CustomizeSimulation = CustomizeSimulation
        self.UpdateProjectParameters = UpdateProjectParameters
        self.ROMvsFOM_train=self.ROMvsHROM_train=self.ROMvsFOM_test=self.ROMvsHROM_test=-1

    def Fit(self, mu_train=[None], mu_train_with_indexes=None, mesh_file=None , update_bases=True):
        mu_train_errors = []
        if (len(mu_train) > 0):
            chosen_projection_strategy = self.general_rom_manager_parameters["projection_strategy"].GetString()
            training_stages = self.general_rom_manager_parameters["rom_stages_to_train"].GetStringArray()

            #######################################
            ##  Least-Squares Petrov Galerkin   ###
            if chosen_projection_strategy == "lspg":
                if any(item == "ROM" for item in training_stages):
                    fom_snapshots = self.__LaunchTrainROM(mu_train_with_indexes, mesh_file, update_bases)
                    rom_snapshots = self.__LaunchROM(mu_train_with_indexes, mesh_file, "lspg")
                    self.ROMvsFOM_train = np.linalg.norm(fom_snapshots - rom_snapshots)/ np.linalg.norm(fom_snapshots)
                    for i in range(len(mu_train)):
                        error = np.linalg.norm(fom_snapshots[:,i] - rom_snapshots[:,i])/ np.linalg.norm(fom_snapshots[:,i])
                        mu_train_errors.append([mu_train[i][0],mu_train[i][1],error])
        return mu_train_errors

    def Test(self, mu_test=[None], mesh_file=None, predicted_indexes=None):
        mu_test_errors = []
        if (len(mu_test) > 0):
            chosen_projection_strategy = self.general_rom_manager_parameters["projection_strategy"].GetString()
            testing_stages = self.general_rom_manager_parameters["rom_stages_to_test"].GetStringArray()

            #######################################
            ##  Least-Squares Petrov Galerkin   ###
            if chosen_projection_strategy == "lspg":
                if any(item == "ROM" for item in testing_stages):
                    fom_snapshots = self.__LaunchTestFOM(mu_test, mesh_file, predicted_indexes)
                    rom_snapshots = self.__LaunchTestROM(mu_test, mesh_file, predicted_indexes, "lspg")
                    self.ROMvsFOM_test = np.linalg.norm(fom_snapshots - rom_snapshots)/ np.linalg.norm(fom_snapshots)
                    for i in range(len(mu_test)):
                        error = np.linalg.norm(fom_snapshots[:,i] - rom_snapshots[:,i])/ np.linalg.norm(fom_snapshots[:,i])
                        mu_test_errors.append([mu_test[i][0],mu_test[i][1],error])
        return mu_test_errors

    def PrintErrors(self):
        training_stages = self.general_rom_manager_parameters["rom_stages_to_train"].GetStringArray()
        testing_stages = self.general_rom_manager_parameters["rom_stages_to_test"].GetStringArray()
        if any(item == "ROM" for item in training_stages) and self.ROMvsFOM_train > 0:
            print("approximation error in train set FOM vs ROM: ", self.ROMvsFOM_train)
        if any(item == "ROM" for item in testing_stages) and self.ROMvsFOM_test > 0:
            print("approximation error in test set FOM vs ROM: ", self.ROMvsFOM_test)
        return 'train', self.ROMvsFOM_train, 'test', self.ROMvsFOM_test

    def __LaunchTrainROM(self, mu_train_with_indexes, mesh_file, update_bases):
        if update_bases or not os.path.exists(f'{mesh_file}/SnapshotsMatrix_FOM.npy'):
            KratosMultiphysics.kratos_utilities.DeleteDirectoryIfExisting(f"{mesh_file}/RomBases")
            os.mkdir(f"{mesh_file}/RomBases")
            with open(self.project_parameters_name,'r') as parameter_file:
                parameters = KratosMultiphysics.Parameters(parameter_file.read())

            n_clusters = np.int0(np.max(np.array(mu_train_with_indexes)[:,2])) + 1
            SnapshotsMatrix = []; NumberOfRomModes_list = []
            
            for i in range(n_clusters):
                SnapshotsMatrix_c = []; SnapshotsMatrix_nc = []
                mu_aux = [[angle, mach] for angle, mach, cluster in mu_train_with_indexes if cluster == i]

                for mu in mu_aux:
                    file = f'{mu[0]}, {mu[1]}.npy'
                    SnapshotsMatrix.append(np.load(f'DataBase/Snapshots/{file}'))
                    SnapshotsMatrix_c.append(np.load(f'DataBase/Snapshots/{file}'))
                    SnapshotsMatrix_nc.append(np.load(f'DataBase/Snapshots_not_converged_steps/{file}'))
                SnapshotsMatrix_c  = np.block(SnapshotsMatrix_c)
                SnapshotsMatrix_nc = np.block(SnapshotsMatrix_nc)

                parameters_copy = self.UpdateProjectParameters(parameters.Clone(), mu, mesh_file)
                parameters_copy = self._AddBasisCreationToProjectParameters(parameters_copy) 
                parameters_copy["output_processes"]["rom_output"][0]["Parameters"]["rom_basis_output_folder"].SetString(f"{mesh_file}/RomBases/rom_data_cluster_"+str(i))
                model = KratosMultiphysics.Model()
                analysis_stage_class = self._GetAnalysisStageClass(parameters_copy)
                simulation = self.CustomizeSimulation(analysis_stage_class, model, parameters_copy, mesh_file, i)
                simulation.Initialize()
                
                for process in simulation._GetListOfOutputProcesses():
                    if isinstance(process, CalculateRomBasisOutputProcess):
                        BasisOutputProcess = process
    
                phi_c,_,_,_= RandomizedSingularValueDecomposition().Calculate(SnapshotsMatrix_c, 1e-12)
                S_nc = SnapshotsMatrix_nc
                for k in range(5):
                    S_nc = S_nc - (phi_c @ (phi_c.T @ S_nc))
                phi_nc,_,_,_= RandomizedSingularValueDecomposition().Calculate(S_nc, 1e-12)

                BasisOutputProcess._PrintRomBasis(np.linalg.svd(np.c_[phi_c, phi_nc], full_matrices=False)[0])

                np.save(f'{mesh_file}/RomBases/rom_data_cluster_{i}/SnapshotsMatrix_c', SnapshotsMatrix_c)
                np.save(f'{mesh_file}/RomBases/rom_data_cluster_{i}/SnapshotsMatrix_nc', SnapshotsMatrix_nc)
                NumberOfRomModes_list.append(np.load(f'{mesh_file}/RomBases/rom_data_cluster_{i}/RightBasisMatrix.npy').shape[1])
                
            SnapshotsMatrix    = np.block(SnapshotsMatrix)
            np.save(f'{mesh_file}/SnapshotsMatrix_FOM', SnapshotsMatrix)
            np.save(f'{mesh_file}/Rom_Modes_List',NumberOfRomModes_list)

            wb = openpyxl.Workbook()
            hoja = wb.active
            hoja.append(['Cluster','Number of modes'])
            for n, item in enumerate(NumberOfRomModes_list):
                hoja.append([n, item])
            wb.save(f'{mesh_file}/rom_modes.xlsx')
        else:
            SnapshotsMatrix = np.load(f'{mesh_file}/SnapshotsMatrix_FOM.npy')

        return SnapshotsMatrix

    def __LaunchROM(self, mu_train_with_indexes, mesh_file, simulation_to_run):
        if not os.path.exists(f"{mesh_file}/ROM_Results"):
            os.mkdir(f"{mesh_file}/ROM_Results")
        if not os.path.exists(f"{mesh_file}/ROM_Skin_Data"):
            os.mkdir(f"{mesh_file}/ROM_Skin_Data")
        if not os.path.exists(f"{mesh_file}/ROM_Captures"):
            os.mkdir(f"{mesh_file}/ROM_Captures")
        else:
            KratosMultiphysics.kratos_utilities.DeleteDirectoryIfExisting(f"{mesh_file}/ROM_Captures")
            os.mkdir(f"{mesh_file}/ROM_Captures")
        KratosMultiphysics.kratos_utilities.DeleteFileIfExisting(f"{mesh_file}/rom_data.xlsx")

        with open(self.project_parameters_name,'r') as parameter_file:
            parameters = KratosMultiphysics.Parameters(parameter_file.read())
            
        n_clusters = np.int0(np.max(np.array(mu_train_with_indexes)[:,2])) + 1
        SnapshotsMatrix = []

        for i in range(n_clusters):
            mu_aux = [[angle, mach] for angle, mach, cluster in mu_train_with_indexes if cluster == i]

            for mu in mu_aux:
                rom_info_steps_list = []
                self.general_rom_manager_parameters["ROM"]["rom_basis_output_folder"].SetString(f'{mesh_file}/RomBases/rom_data_cluster_{i}')
                self._ChangeRomFlags(simulation_to_run)
                parameters_copy = self.UpdateProjectParameters(parameters.Clone(), mu, mesh_file)
                parameters_copy = self._AddBasisCreationToProjectParameters(parameters_copy)  
                model = KratosMultiphysics.Model()
                analysis_stage_class = type(SetUpSimulationInstance(model, parameters_copy))
                simulation = self.CustomizeSimulation(analysis_stage_class, model, parameters_copy, mesh_file, i)
                simulation.Run()
                for process in simulation._GetListOfOutputProcesses():
                    if isinstance(process, CalculateRomBasisOutputProcess):
                        BasisOutputProcess = process
                SnapshotsMatrix.append(BasisOutputProcess._GetSnapshotsMatrix())
                
                model_part = simulation.model[parameters_copy["solver_settings"]["model_part_name"].GetString()].GetRootModelPart()
                file = f'{mu[0]}, {mu[1]}.npy'
                fom  = np.load(f'DataBase/Snapshots/{file}')
                rom  = BasisOutputProcess._GetSnapshotsMatrix()
                rom_info_steps_list.append([i,
                                            mu[0],
                                            mu[1], 
                                            model_part.ProcessInfo[KratosMultiphysics.NL_ITERATION_NUMBER],
                                            model_part.ProcessInfo[KratosMultiphysics.RESIDUAL_NORM],
                                            np.linalg.norm(fom-rom)/np.linalg.norm(fom),
                                            ])

                if os.path.exists(f"{mesh_file}/rom_data.xlsx"):
                    wb = openpyxl.load_workbook(f"{mesh_file}/rom_data.xlsx")
                    hoja = wb.active
                    for item in rom_info_steps_list:
                        hoja.append(item)
                    wb.save(f'{mesh_file}/rom_data.xlsx')
                else:
                    wb = openpyxl.Workbook()
                    hoja = wb.active
                    hoja.append(('Cluster', 'Angle [º]', 'Mach', 'NL iterations', 'Residual norm', 'Approximation error'))
                    for item in rom_info_steps_list:
                        hoja.append(item)
                    wb.save(f'{mesh_file}/rom_data.xlsx')

        SnapshotsMatrix = np.block(SnapshotsMatrix)
        np.save(f'{mesh_file}/SnapshotsMatrix_ROM', SnapshotsMatrix)
        return SnapshotsMatrix

    def __LaunchTestFOM(self, mu_test, mesh_file, predicted_indexes):
        if not os.path.exists(f"{mesh_file}/Test"):
            os.mkdir(f"{mesh_file}/Test")
        n_clusters = np.int0(np.max(np.array(predicted_indexes))) + 1
        SnapshotsMatrix = []

        for i in range(n_clusters):
            mu_aux = [[angle, mach] for (angle, mach), cluster in zip(mu_test, predicted_indexes) if cluster == i]

            for mu in mu_aux:
                file = f'{mu[0]}, {mu[1]}.npy'
                SnapshotsMatrix.append(np.load(f'DataBase/Snapshots/{file}'))

        SnapshotsMatrix    = np.block(SnapshotsMatrix)
        np.save(f'{mesh_file}/Test/SnapshotsMatrix_FOM_Test', SnapshotsMatrix)
        return SnapshotsMatrix


    def __LaunchTestROM(self, mu_test, mesh_file, predicted_indexes, simulation_to_run):
        if not os.path.exists(f"{mesh_file}/Test/ROM_Results"):
            os.mkdir(f"{mesh_file}/Test/ROM_Results")
        if not os.path.exists(f"{mesh_file}/Test/ROM_Skin_Data"):
            os.mkdir(f"{mesh_file}/Test/ROM_Skin_Data")
        if not os.path.exists(f"{mesh_file}/Test/ROM_Captures"):
            os.mkdir(f"{mesh_file}/Test/ROM_Captures")
        else:
            KratosMultiphysics.kratos_utilities.DeleteDirectoryIfExisting(f"{mesh_file}/Test/ROM_Captures")
            os.mkdir(f"{mesh_file}/Test/ROM_Captures")
        KratosMultiphysics.kratos_utilities.DeleteFileIfExisting(f"{mesh_file}/Test/rom_data.xlsx")

        with open(self.project_parameters_name,'r') as parameter_file:
            parameters = KratosMultiphysics.Parameters(parameter_file.read())
        n_clusters = np.int0(np.max(np.array(predicted_indexes))) + 1
        SnapshotsMatrix = []

        for i in range(n_clusters):
            mu_aux = [[angle, mach] for (angle, mach), cluster in zip(mu_test, predicted_indexes) if cluster == i]

            for mu in mu_aux:
                rom_info_steps_list = []
                self.general_rom_manager_parameters["ROM"]["rom_basis_output_folder"].SetString(f'{mesh_file}/RomBases/rom_data_cluster_{i}')
                self._ChangeRomFlags(simulation_to_run)
                parameters_copy = self.UpdateProjectParameters(parameters.Clone(), mu, mesh_file, True)
                parameters_copy = self._AddBasisCreationToProjectParameters(parameters_copy)  
                model = KratosMultiphysics.Model()
                analysis_stage_class = type(SetUpSimulationInstance(model, parameters_copy))
                simulation = self.CustomizeSimulation(analysis_stage_class, model, parameters_copy, mesh_file, i, True)
                simulation.Run()
                for process in simulation._GetListOfOutputProcesses():
                    if isinstance(process, CalculateRomBasisOutputProcess):
                        BasisOutputProcess = process
                SnapshotsMatrix.append(BasisOutputProcess._GetSnapshotsMatrix()) 
                
                model_part = simulation.model[parameters_copy["solver_settings"]["model_part_name"].GetString()].GetRootModelPart()
                file = f'{mu[0]}, {mu[1]}.npy'
                fom  = np.load(f'DataBase/Snapshots/{file}')
                rom  = BasisOutputProcess._GetSnapshotsMatrix()
                rom_info_steps_list.append([i,
                                            mu[0],
                                            mu[1], 
                                            model_part.ProcessInfo[KratosMultiphysics.NL_ITERATION_NUMBER],
                                            model_part.ProcessInfo[KratosMultiphysics.RESIDUAL_NORM],
                                            np.linalg.norm(fom-rom)/np.linalg.norm(fom),
                                            ])
                
                if os.path.exists(f"{mesh_file}/Test/rom_data.xlsx"):
                    wb = openpyxl.load_workbook(f"{mesh_file}/Test/rom_data.xlsx")
                    hoja = wb.active
                    for item in rom_info_steps_list:
                        hoja.append(item)
                    wb.save(f'{mesh_file}/Test/rom_data.xlsx')
                else:
                    wb = openpyxl.Workbook()
                    hoja = wb.active
                    hoja.append(('Cluster', 'Angle', 'Mach', 'NL iterations', 'Residual norm', 'Approximation error'))
                    for item in rom_info_steps_list:
                        hoja.append(item)
                    wb.save(f'{mesh_file}/Test/rom_data.xlsx')

        SnapshotsMatrix = np.block(SnapshotsMatrix)
        np.save(f'{mesh_file}/Test/SnapshotsMatrix_ROM_Test', SnapshotsMatrix)
        return SnapshotsMatrix

    def _ChangeRomFlags(self, simulation_to_run = 'ROM'):
        """
        This method updates the Flags present in the RomParameters.json file
        for launching the correct part of the ROM workflow
        """
        parameters_file_folder = self.general_rom_manager_parameters["ROM"]["rom_basis_output_folder"].GetString() if self.general_rom_manager_parameters["ROM"].Has("rom_basis_output_folder") else "rom_data"
        parameters_file_name = self.general_rom_manager_parameters["ROM"]["rom_basis_output_name"].GetString() if self.general_rom_manager_parameters["ROM"].Has("rom_basis_output_name") else "RomParameters"

        parameters_file_folder = Path(parameters_file_folder)
        parameters_file_name = Path(parameters_file_name)

        parameters_file_path = parameters_file_folder / parameters_file_name.with_suffix('.json')

        with parameters_file_path.open('r+') as parameter_file:
            f=json.load(parameter_file)
            f['assembling_strategy'] = self.general_rom_manager_parameters['assembling_strategy'].GetString() if self.general_rom_manager_parameters.Has('assembling_strategy') else 'global'
            if simulation_to_run == 'lspg':
                f['train_hrom'] = False
                f['run_hrom'] = False
                f['projection_strategy'] = "lspg"
                f["rom_settings"]['rom_bns_settings'] = self._SetLSPGBnSParameters()
            else:
                raise Exception(f'Unknown flag "{simulation_to_run}" change for RomParameters.json')
            parameter_file.seek(0)
            json.dump(f,parameter_file,indent=4)
            parameter_file.truncate()

    def _SetLSPGBnSParameters(self):
        defaults_json = self._GetLSPGBnSParameters()
        defaults = json.loads(defaults_json)

        if not self.general_rom_manager_parameters["ROM"].Has("lspg_rom_bns_settings"):
            self.general_rom_manager_parameters["ROM"].AddEmptyValue("lspg_rom_bns_settings")

        rom_params = self.general_rom_manager_parameters["ROM"]["lspg_rom_bns_settings"]

        self._UpdateDefaultsWithRomParams(defaults, rom_params)

        return defaults

    def _UpdateDefaultsWithRomParams(self, defaults, rom_params):
        for key, default_value in defaults.items():
            if rom_params.Has(key):
                if isinstance(default_value, bool):
                    defaults[key] = rom_params[key].GetBool()
                elif isinstance(default_value, str):
                    defaults[key] = rom_params[key].GetString()
                elif isinstance(default_value, float):
                    defaults[key] = rom_params[key].GetDouble()
        return defaults

    def _AddBasisCreationToProjectParameters(self, parameters):
        parameters["output_processes"].AddEmptyArray("rom_output")
        parameters["output_processes"]["rom_output"].Append(self._SetRomTrainingParameters())

        return parameters

    def _StoreNoResults(self, parameters):
        parameters["output_processes"].RemoveValue("gid_output")
        parameters["output_processes"].RemoveValue("vtk_output")
        return parameters

    def _SetRomTrainingParameters(self):
        defaults = self._GetDefaulRomBasisOutputParameters()
        defaults["Parameters"]["rom_manager"].SetBool(True)  
        rom_params = self.general_rom_manager_parameters["ROM"]
        keys_to_copy = [
            "svd_truncation_tolerance",
            "model_part_name",
            "rom_basis_output_format",
            "rom_basis_output_name",
            "rom_basis_output_folder",
            "nodal_unknowns",
            "snapshots_interval",
        ]
        for key in keys_to_copy:
            if key in rom_params.keys():
                defaults["Parameters"][key] = rom_params[key]
        return defaults

    def _GetAnalysisStageClass(self, parameters):
        analysis_stage_module_name = parameters["analysis_stage"].GetString()
        analysis_stage_class_name = analysis_stage_module_name.split('.')[-1]
        analysis_stage_class_name = ''.join(x.title() for x in analysis_stage_class_name.split('_'))
        analysis_stage_module = importlib.import_module(analysis_stage_module_name)
        analysis_stage_class = getattr(analysis_stage_module, analysis_stage_class_name)
        return analysis_stage_class

    def _GetDefaulRomBasisOutputParameters(self):
        rom_training_parameters = KratosMultiphysics.Parameters("""{
                "python_module" : "calculate_rom_basis_output_process",
                "kratos_module" : "KratosMultiphysics.RomApplication",
                "process_name"  : "CalculateRomBasisOutputProcess",
                "help"          : "This process should write the Rom basis",
                "Parameters"    :
                {
                    "model_part_name": "",
                    "rom_manager" : false,      // set to false for manual manipulation of ROM via flags in the RomParameters
                    "snapshots_control_type": "step",
                    "snapshots_interval": 1.0,
                    "nodal_unknowns":  [],
                    "rom_basis_output_format": "json",
                    "rom_basis_output_name": "RomParameters",
                    "rom_basis_output_folder": "rom_data",
                    "svd_truncation_tolerance": 1e-3
                }
            }""")
        return rom_training_parameters

    def _GetLSPGBnSParameters(self):
        rom_bns_settings = """{
            "train_petrov_galerkin": false,
            "basis_strategy": "residuals",
            "include_phi": false,
            "svd_truncation_tolerance": 1e-8,
            "solving_technique": "normal_equations",
            "monotonicity_preserving": false
        }"""
        return rom_bns_settings