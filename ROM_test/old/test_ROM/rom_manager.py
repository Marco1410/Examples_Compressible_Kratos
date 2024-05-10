import os
import openpyxl
import KratosMultiphysics.kratos_utilities
import numpy as np
import matplotlib
import matplotlib.pyplot as plt
matplotlib.use('Agg')
import KratosMultiphysics
from local_rom_manager import RomManager # type: ignore
from PEBL_RBF import PEBL_RBF_Clustering # type: ignore
import KratosMultiphysics.CompressiblePotentialFlowApplication as CPFApp 
from KratosMultiphysics.RomApplication.calculate_rom_basis_output_process import CalculateRomBasisOutputProcess


# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #


def CustomizeSimulation(cls, global_model, parameters):

    class CustomSimulation(cls):

        def __init__(self, model,project_parameters):
            super().__init__(model,project_parameters)
            """
            Customize as needed
            """
            self.simulation_name = 'Run'
            self.cluster = -1
            self.modes = -1

            if parameters["output_processes"].Has("gid_output"):
                if not os.path.exists('Skin_Data'):
                    os.mkdir('Skin_Data')
                if not os.path.exists(f'Captures'):
                    os.mkdir('Captures')
                self.simulation_name = parameters["output_processes"]["gid_output"][0]["Parameters"]["output_name"].GetString().removeprefix('Results/')

            self.angle = parameters["processes"]["boundary_conditions_process_list"][0]["Parameters"]["angle_of_attack"].GetDouble()
            self.mach  = parameters["processes"]["boundary_conditions_process_list"][0]["Parameters"]["mach_infinity"].GetDouble()

            self.fom_name = f'{self.angle}, {self.mach}'
            self.save_solution = False

            if os.path.exists(f'mu_train_with_indexes.npy') and os.path.exists('predicted_indexes.npy'):
                mu_train = np.load('mu_train.npy')
                mu_test  = np.load('mu_test.npy')
                finded = False
                if not finded:
                    for i in range(len(mu_train)):
                        if mu_train[i][0] == self.angle and mu_train[i][1] == self.mach:
                            self.cluster = np.int0(np.load('mu_train_with_indexes.npy')[i][2])
                            finded = True
                            break
                if not finded:
                    for i in range(len(mu_test)):
                        if mu_test[i][0] == self.angle and mu_test[i][1] == self.mach:
                            self.cluster = np.int0(np.load('predicted_indexes.npy')[i])
                            finded = True
                            break

        def Initialize(self):
            super().Initialize()
            """
            Customize as needed
            """
            # set_initialize_solution = False
            # if set_initialize_solution and self._GetSimulationName() == "::[ROM Simulation]:: ":
            #     u = np.load(f'DataBase/Snapshots/{self.fom_name}.npy')    # The converged solution
                # u = np.load(f'Centroids/Centroid_{self.cluster}.npy').T # The centroid of the cluster
                # phi = np.load(f'rom_bases/rom_data_cluster_{self.cluster}/RightBasisMatrix.npy')
                # u = phi@(phi.T@u)
                # model_part = self.model["MainModelPart"]
                # for node in model_part.Nodes:
                #     offset = np.where(np.arange(1,model_part.NumberOfNodes()+1, dtype=int) == node.Id)[0][0]*2
                #     node.SetSolutionStepValue(CPFApp.VELOCITY_POTENTIAL, u[offset+1])
                #     node.SetSolutionStepValue(CPFApp.AUXILIARY_VELOCITY_POTENTIAL, u[offset])

        def Run(self):
            """
            Customize as needed
            """
            info_steps_list = []
            if self._GetSimulationName() == "Analysis": # FOM
                if not os.path.exists('DataBase'):
                    os.mkdir('DataBase')
                if not os.path.exists('DataBase/Snapshots'):
                    os.mkdir('DataBase/Snapshots')
                if not os.path.exists('DataBase/Snapshots_not_converged/'):
                    os.mkdir('DataBase/Snapshots_not_converged/')
                self.Initialize()
                if not os.path.exists(f'DataBase/Snapshots/{self.fom_name}.npy'):
                    self.RunSolutionLoop()
                    self.save_solution = True
                else:
                    values = np.load(f'DataBase/Snapshots/{self.fom_name}.npy')
                    for node in self.model["MainModelPart"].Nodes:
                        offset = np.where(np.arange(1,self.model["MainModelPart"].NumberOfNodes()+1, dtype=int) == node.Id)[0][0]*2
                        node.SetSolutionStepValue(CPFApp.VELOCITY_POTENTIAL, values[offset+1])
                        node.SetSolutionStepValue(CPFApp.AUXILIARY_VELOCITY_POTENTIAL, values[offset])

                    while self.KeepAdvancingSolutionLoop():
                        self.time = self._AdvanceTime()
                        self.InitializeSolutionStep()
                        self.FinalizeSolutionStep()
                        self.OutputSolutionStep()
                self.Finalize()

                if self.save_solution:
                    for process in self._GetListOfOutputProcesses():
                        if isinstance(process, CalculateRomBasisOutputProcess):
                            BasisOutputProcess = process
                    np.save(f'DataBase/Snapshots/{self.fom_name}',BasisOutputProcess._GetSnapshotsMatrix()) 
                    np.save(f'DataBase/Snapshots_not_converged/{self.fom_name}', np.array(self._GetSolver()._GetSolutionStrategy().GetIntermediateSolutionsMatrix()))

                if self.cluster >= 0:
                    if os.path.exists(f'rom_bases/rom_data_cluster_{self.cluster}/RightBasisMatrix.npy'):
                        self.modes = np.load(f'rom_bases/rom_data_cluster_{self.cluster}/RightBasisMatrix.npy').shape[1]
                info_steps_list.append([self.cluster,
                                        self.angle,
                                        self.mach, 
                                        self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.NL_ITERATION_NUMBER],
                                        self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.RESIDUAL_NORM],
                                        self.modes,
                                        ])
                if os.path.exists(f"fom_data.xlsx"):
                    wb = openpyxl.load_workbook(f"fom_data.xlsx")
                    hoja = wb.active
                    for item in info_steps_list:
                        hoja.append(item)
                    wb.save(f'fom_data.xlsx')
                else:
                    wb = openpyxl.Workbook()
                    hoja = wb.active
                    hoja.append(('Cluster', 'Angle [º]', 'Mach', 'NL iterations', 'Residual norm', 'Modes'))
                    for item in info_steps_list:
                        hoja.append(item)
                    wb.save(f'fom_data.xlsx')

            else:

                if not os.path.exists('ROM_Solutions'):
                    os.mkdir('ROM_Solutions')
                self.Initialize()
                self.RunSolutionLoop()
                self.Finalize()
                for process in self._GetListOfOutputProcesses():
                        if isinstance(process, CalculateRomBasisOutputProcess):
                            BasisOutputProcess = process
                fom = np.load(f'DataBase/Snapshots/{self.fom_name}.npy')
                rom = BasisOutputProcess._GetSnapshotsMatrix()

                if np.linalg.norm(fom-rom)/np.linalg.norm(fom) > 1e-2:
                    np.save(f'ROM_Solutions/{self.fom_name}',BasisOutputProcess._GetSnapshotsMatrix()) 
            
                info_steps_list.append([self.cluster,
                                        self.angle,
                                        self.mach, 
                                        self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.NL_ITERATION_NUMBER],
                                        self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.RESIDUAL_NORM],
                                        np.linalg.norm(fom-rom)/np.linalg.norm(fom),
                                        ])
                if os.path.exists(f"rom_data.xlsx"):
                    wb = openpyxl.load_workbook(f"rom_data.xlsx")
                    hoja = wb.active
                    for item in info_steps_list:
                        hoja.append(item)
                    wb.save(f'rom_data.xlsx')
                else:
                    wb = openpyxl.Workbook()
                    hoja = wb.active
                    hoja.append(('Cluster', 'Angle [º]', 'Mach', 'NL iterations', 'Residual norm', 'Approximation error [%]'))
                    for item in info_steps_list:
                        hoja.append(item)
                    wb.save(f'rom_data.xlsx')

        def FinalizeSolutionStep(self):
            super().FinalizeSolutionStep()
            """
            Customize as needed
            """
            if 'Run' not in self.simulation_name:
                if parameters["output_processes"].Has("gid_output"):

                    if not os.path.exists(f'Captures/Cluster_{self.cluster}'):
                        os.mkdir(f'Captures/Cluster_{self.cluster}')

                    skin_data_filename = f"Skin_Data/{self.simulation_name}.dat"
                    capture_filename   = f"Captures/Cluster_{self.cluster}/{self.simulation_name}.png"

                    fout = open(skin_data_filename,'w')
                    modelpart = self.model["MainModelPart.Body2D_Body"]
                    for node in modelpart.Nodes:
                        x = node.X ; y = node.Y ; z = node.Z
                        cp = node.GetValue(KratosMultiphysics.PRESSURE_COEFFICIENT)
                        fout.write("%s %s %s %s\n" %(x,y,z,cp))
                    fout.close()

                    #### ONLINE CP PLOT
                    ######################################################################
                    cp_min = 0
                    cp_max = 0
                    fig = plt.figure()
                    fig.set_figwidth(12.0)
                    fig.set_figheight(8.0)
                    x  = np.loadtxt(skin_data_filename, usecols=(0,))
                    cp = np.loadtxt(skin_data_filename, usecols=(3,))
                    fig = plt.plot(x, cp, 'xr', markersize = 2.0, label = self.simulation_name)
                    if np.min(cp) < cp_min:
                        cp_min = np.min(cp)
                    if np.max(cp) > cp_max:
                        cp_max = np.max(cp)
                    fig = plt.title('Cp vs x')
                    fig = plt.axis([-0.05,1.35,cp_max+0.1,cp_min-0.1])
                    fig = plt.ylabel('Cp')
                    fig = plt.xlabel('x')
                    fig = plt.grid()
                    fig = plt.legend()
                    fig = plt.tight_layout()
                    fig = plt.savefig(capture_filename)
                    fig = plt.close('all')

    return CustomSimulation(global_model, parameters)


# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

def UpdateProjectParameters(parameters, mu=None):
    """
    Customize ProjectParameters here for imposing different conditions to the simulations as needed
    """
    angle_of_attack        = mu[0]
    mach_infinity          = mu[1]
    parameters["processes"]["boundary_conditions_process_list"][0]["Parameters"]["angle_of_attack"].SetDouble(angle_of_attack)
    parameters["processes"]["boundary_conditions_process_list"][0]["Parameters"]["mach_infinity"].SetDouble(mach_infinity)
    return parameters

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

def UpdateMaterialParametersFile(material_parametrs_file_name, mu):
    pass
    
# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
#
# get multiple parameters by Halton or LatinHypercube methods
#
def get_multiple_parameters(number_train_values=0, number_test_values=0, angle=[], mach=[], method='Halton'):
    from scipy.stats import qmc                 
    if method == 'Halton':
        sampler = qmc.Halton(d=2)
    elif method == 'LatinHypercube':
        sampler = qmc.LatinHypercube(d=2)
    mu_train = []; mu_test = []; aux_mu_train = []; aux_mu_test = []
    if number_train_values > 0:
        sample = sampler.random(number_train_values)
        values = qmc.scale(sample, [angle[0],mach[0]], [angle[1],mach[1]])
        for i in range(number_train_values):
            #Angle of attack , Mach infinit
            mu_train.append([values[i,0], values[i,1]])
            aux_mu_train.append([sample[i,0], sample[i,1]])
    if number_test_values > 0:
        sample = sampler.random(number_test_values)
        values = qmc.scale(sample, [angle[0],mach[0]], [angle[1],mach[1]])
        for i in range(number_test_values):
            #Angle of attack , Mach infinit
            mu_test.append([values[i,0], values[i,1]])
            aux_mu_test.append([sample[i,0], sample[i,1]])
    return mu_train, mu_test, aux_mu_train, aux_mu_test

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
#
# load parameters
#
def load_mu_parameters():
    if os.path.exists("mu_train.npy") and os.path.exists("mu_test.npy"):
        mu_train = np.load('mu_train.npy')
        aux_mu_train = np.load('aux_mu_train.npy')
        mu_test = np.load('mu_test.npy')
        aux_mu_test = np.load('aux_mu_test.npy')
    elif os.path.exists("mu_train.npy"):
        mu_train = np.load('mu_train.npy')
        aux_mu_train = np.load('aux_mu_train.npy')
        mu_test = []
        aux_mu_test = []
    elif os.path.exists("mu_test.npy"):
        mu_test = np.load('mu_test.npy')
        aux_mu_test = np.load('aux_mu_test.npy')
        mu_train = []
        aux_mu_train = []
    return np.array(mu_train), np.array(mu_test), aux_mu_train, aux_mu_test

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
#
# plot parameters
#
def plot_mu_values(mu_train, mu_test):
    if len(mu_train) > 0: plt.plot(mu_train[:,1], mu_train[:,0], 'bs', label="Train Values")
    if len(mu_test) > 0: plt.plot(mu_test[:,1], mu_test[:,0], 'ro', label="Test Values")
    plt.title('Mu Values')
    plt.ylabel('Alpha')
    plt.xlabel('Mach')
    plt.grid(True)
    plt.legend(bbox_to_anchor=(.85, 1.03, 1., .102), loc='upper left', borderaxespad=0.)
    plt.savefig("MuValues.png")
    plt.close('all')

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
#
# PEBL clustering
# 
def error_estimation(mu_train, mu_test):
    if len(mu_train) > 0:
        approximation_error = 0.0
        FOM_model = []; RBF_model = []
        for mu in mu_train:
            FOM_model.append(np.load(f'DataBase/Snapshots/{mu[0]}, {mu[1]}.npy'))
            RBF_model.append(np.load(f"Interpolated_data/{mu[0]}, {mu[1]}.npy").T)
        FOM_model = np.block(FOM_model)
        RBF_model = np.block(RBF_model)
        training_approximation_error = np.linalg.norm(FOM_model - RBF_model)/np.linalg.norm(FOM_model)
        print(f'PEBL CLUSTERING: RBF training approximation error: {training_approximation_error:.2E}')

    if len(mu_test) > 0:
        approximation_error = 0.0
        FOM_model = []; RBF_model_interpolation = []
        for mu in mu_test:
            FOM_model.append(np.load(f'DataBase/Snapshots/{mu[0]}, {mu[1]}.npy'))
            RBF_model_interpolation.append(np.load(f"Interpolated_data/{mu[0]}, {mu[1]}.npy").T)
        FOM_model = np.block(FOM_model)
        RBF_model_interpolation = np.block(RBF_model_interpolation)
        approximation_error = np.linalg.norm(FOM_model - RBF_model_interpolation)/np.linalg.norm(FOM_model)
        print(f'PEBL CLUSTERING: RBF interpolation approximation error: {approximation_error:.2E}')

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #


def GetRomManagerParameters():
    """
    This function allows to easily modify all the parameters for the ROM simulation.
    The returned KratosParameter object is seamlessly used inside the RomManager.
    """
    general_rom_manager_parameters = KratosMultiphysics.Parameters("""{
            "rom_stages_to_train" : [],             // ["ROM","HROM"]
            "rom_stages_to_test"  : [],             // ["ROM","HROM"]
            "paralellism" : null,                        // null, TODO: add "compss"
            "projection_strategy": "lspg",            // "lspg", "galerkin", "petrov_galerkin"
            "assembling_strategy": "global",            // "global", "elemental"
            "save_gid_output": true,                    // false, true #if true, it must exits previously in the ProjectParameters.json
            "save_vtk_output": false,                    // false, true #if true, it must exits previously in the ProjectParameters.json
            "output_name": "mu",                         // "id" , "mu"
            "ROM":{
                "svd_truncation_tolerance": 1e-12,
                "model_part_name": "MainModelPart",                            // This changes depending on the simulation: Structure, FluidModelPart, ThermalPart #TODO: Idenfity it automatically
                "nodal_unknowns": ["VELOCITY_POTENTIAL","AUXILIARY_VELOCITY_POTENTIAL"],     // Main unknowns. Snapshots are taken from these
                "rom_basis_output_format": "numpy",
                "rom_basis_output_name": "RomParameters",
                "rom_basis_output_folder": "rom_data",
                "snapshots_control_type": "step",                          // "step", "time"
                "snapshots_interval": 1,
                "galerkin_rom_bns_settings": {
                    "monotonicity_preserving": false
                },
                "lspg_rom_bns_settings": {
                    "train_petrov_galerkin": false,
                    "basis_strategy": "reactions",                        // 'residuals', 'jacobian', 'reactions'
                    "include_phi": false,
                    "svd_truncation_tolerance": 1e-12,
                    "solving_technique": "normal_equations",              // 'normal_equations', 'qr_decomposition'
                    "monotonicity_preserving": false
                },
                "petrov_galerkin_rom_bns_settings": {
                    "monotonicity_preserving": false
                }
            },
            "HROM":{
                "element_selection_type": "empirical_cubature",
                "element_selection_svd_truncation_tolerance": 1e-12,
                "create_hrom_visualization_model_part" : false,
                "echo_level" : 0
            }
        }""")

    return general_rom_manager_parameters


# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #


if __name__ == "__main__":

    update_parameters        = True
    Launch_Fom_Simulations   = True
    update_clusters          = True
    plot_custering_info      = True
    plot_clusters            = True
    plot_clusters_prediction = True
    set_train_rom            = True
    set_test_rom             = True

    KratosMultiphysics.kratos_utilities.DeleteDirectoryIfExisting('Captures')
    KratosMultiphysics.kratos_utilities.DeleteDirectoryIfExisting('Results')
    KratosMultiphysics.kratos_utilities.DeleteDirectoryIfExisting('Skin_Data')

    ###############################
    # PARAMETERS SETTINGS
    number_of_mu_train = 100
    number_of_mu_test  = 100
    mach_range         = [ 0.70, 0.75]
    angle_range        = [ 1.00, 2.00]
    ###############################
    # CLUSTERING SETTINGS
    bisection_tolerance = 0.10
    POD_tolerance       = 1e-12
    ###############################
    
    if update_parameters:
        KratosMultiphysics.kratos_utilities.DeleteDirectoryIfExisting('DataBase')
        KratosMultiphysics.kratos_utilities.DeleteFileIfExisting('mu_train.npy')
        KratosMultiphysics.kratos_utilities.DeleteFileIfExisting('mu_test.npy')
        KratosMultiphysics.kratos_utilities.DeleteFileIfExisting('MuValues.png')
        KratosMultiphysics.kratos_utilities.DeleteFileIfExisting('train_errors.npy')
        KratosMultiphysics.kratos_utilities.DeleteFileIfExisting('test_errors.npy')

        mu_train, mu_test, aux_mu_train, aux_mu_test = get_multiple_parameters(
                                                                                number_train_values = number_of_mu_train,
                                                                                number_test_values  = number_of_mu_test , 
                                                                                angle               = angle_range       , 
                                                                                mach                = mach_range        , 
                                                                                method              = 'Halton'  )
    else:
        mu_train, mu_test, aux_mu_train, aux_mu_test = load_mu_parameters()
    
    plot_mu_values(np.array(mu_train), np.array(mu_test))

    np.save(f'mu_train', mu_train); np.save(f'mu_test', mu_test)
    np.save(f'aux_mu_train', aux_mu_train); np.save(f'aux_mu_test', aux_mu_test)

    #::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::
    if Launch_Fom_Simulations:
        general_rom_manager_parameters = GetRomManagerParameters()

        project_parameters_name = "ProjectParameters.json"

        rom_manager = RomManager(   project_parameters_name, 
                                    general_rom_manager_parameters, 
                                    CustomizeSimulation, 
                                    UpdateProjectParameters, 
                                    UpdateMaterialParametersFile )
            
        rom_manager.RunFOM(mu_run = mu_train) 
        KratosMultiphysics.kratos_utilities.DeleteDirectoryIfExisting('rom_data')
    #::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::
    
    if update_parameters: update_clusters = True
    if update_clusters:
        KratosMultiphysics.kratos_utilities.DeleteDirectoryIfExisting('Clustering_mu_test_plots')
        KratosMultiphysics.kratos_utilities.DeleteDirectoryIfExisting('Interpolated_data')
        KratosMultiphysics.kratos_utilities.DeleteDirectoryIfExisting('rom_bases')
        KratosMultiphysics.kratos_utilities.DeleteDirectoryIfExisting('rom_data')
        KratosMultiphysics.kratos_utilities.DeleteDirectoryIfExisting('Centroids')
        KratosMultiphysics.kratos_utilities.DeleteFileIfExisting('mu_train_with_indexes.npy')
        KratosMultiphysics.kratos_utilities.DeleteFileIfExisting('predicted_indexes.npy')
        KratosMultiphysics.kratos_utilities.DeleteFileIfExisting('PEBL clustering.png')
        KratosMultiphysics.kratos_utilities.DeleteFileIfExisting('leaves.npy')

        leaves, mu_train_with_leaves, predicted_indexes_list = PEBL_RBF_Clustering( update_clusters          = update_clusters,
                                                                                    plot_custering_info      = plot_custering_info,
                                                                                    plot_clusters            = plot_clusters,
                                                                                    plot_clusters_prediction = plot_clusters_prediction,
                                                                                    bisection_tolerance      = bisection_tolerance,
                                                                                    POD_tolerance            = POD_tolerance,
                                                                                    mu_train                 = mu_train,
                                                                                    mu_test                  = mu_test,
                                                                                    aux_mu_train             = aux_mu_train,
                                                                                    aux_mu_test              = aux_mu_test)
    else:
        mu_train_with_leaves    = np.load('mu_train_with_indexes.npy')
        predicted_indexes_list  = np.load('predicted_indexes.npy')
        leaves                  = np.load('leaves.npy', allow_pickle=True)

    train_errors = []; test_errors = []
    if set_train_rom:
        for i in range(len(leaves)):
            mu_aux = [[angle, mach] for angle, mach, cluster in mu_train_with_leaves if cluster == i]

            #::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::
            general_rom_manager_parameters = GetRomManagerParameters()
            general_rom_manager_parameters['rom_stages_to_train'].SetStringArray(['ROM'])
            general_rom_manager_parameters['ROM']['rom_basis_output_folder'].SetString(f'rom_bases/rom_data_cluster_{i}')

            project_parameters_name = "ProjectParameters.json"

            rom_manager = RomManager(   project_parameters_name, 
                                        general_rom_manager_parameters, 
                                        CustomizeSimulation, 
                                        UpdateProjectParameters, 
                                        UpdateMaterialParametersFile )            
            #::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::

            train_error,_ = rom_manager.Fit(mu_aux, store_all_snapshots = True)
            modes = np.load(f'rom_bases/rom_data_cluster_{i}/RightBasisMatrix.npy').shape[1]
            train_errors.append([f'Train error FOM - ROM, Cluster {i}: {train_error:.2E}, Modes: {modes}, Points: {len(mu_aux)}'])
    
    if set_test_rom:
        KratosMultiphysics.kratos_utilities.DeleteDirectoryIfExisting('ROM_Solutions')
        for i in range(len(leaves)):
            mu_aux = [[angle, mach] for (angle, mach), cluster in zip(mu_test, predicted_indexes_list) if cluster == i]

            #::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::
            general_rom_manager_parameters = GetRomManagerParameters()
            general_rom_manager_parameters['rom_stages_to_test'].SetStringArray(['ROM'])
            general_rom_manager_parameters['ROM']['rom_basis_output_folder'].SetString(f'rom_bases/rom_data_cluster_{i}')

            project_parameters_name = "ProjectParameters.json"

            rom_manager = RomManager(   project_parameters_name, 
                                        general_rom_manager_parameters, 
                                        CustomizeSimulation, 
                                        UpdateProjectParameters, 
                                        UpdateMaterialParametersFile )            
            #::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::

            test_error,_ = rom_manager.Test(mu_aux, store_all_snapshots = True)

            test_errors.append([f'Test  error FOM - ROM, cluster {i}: {test_error:.2E}, Points: {len(mu_aux)}'])
            
    if plot_custering_info: 
        print('::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::')
        error_estimation(mu_train, mu_test)

    print('::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::')
    for i in range(len(train_errors)):
        print(train_errors[i])  
    print('::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::')
    for i in range(len(test_errors)):
        print(test_errors[i])
    print('::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::')
    
    np.save('train_errors', train_errors)
    np.save('test_errors', test_errors)
