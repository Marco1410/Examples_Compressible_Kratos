import os
import time
import openpyxl
import numpy as np
import matplotlib
import matplotlib.pyplot as plt
matplotlib.use('Agg')
from scipy.stats import qmc  
import KratosMultiphysics
import KratosMultiphysics.kratos_utilities
import KratosMultiphysics.CompressiblePotentialFlowApplication as CPFApp
from KratosMultiphysics.RomApplication.rom_manager import RomManager
from KratosMultiphysics.RomApplication.calculate_rom_basis_output_process import CalculateRomBasisOutputProcess


# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
#
# get multiple parameters by Halton or LatinHypercube methods
#
def get_multiple_parameters(regions=[], number_test_values=0,
                            angle=[], mach=[], mu_validation=[], 
                            method='Halton', update_mu_test = True, alpha=1.0, beta=1.0):
    if method == 'Halton':
        sampler = qmc.Halton(d=2, scramble=False)
    elif method == 'LatinHypercube':
        sampler = qmc.LatinHypercube(d=2)
    mu_train = []; mu_test = []
    mu_train_not_scaled = []; mu_test_not_scaled = []; mu_validation_not_scaled = []
    for mach_range, alpha_range, number_of_values in regions:
        mach_common_min  = max(mach_range[0] , mach[0])
        mach_common_max  = min(mach_range[1] , mach[1])
        angle_common_min = max(alpha_range[0], angle[0])
        angle_common_max = min(alpha_range[1], angle[1])
        if mach_common_min < mach_common_max and angle_common_min < angle_common_max and number_of_values > 0:
            sample = sampler.random(number_of_values)
            # Density transformation: Alpha: density X; Beta:  density Y
            transformed_points = np.zeros_like(sample)
            transformed_points[:, 0] = sample[:, 0]**beta
            transformed_points[:, 1] = sample[:, 1]**alpha
            values = qmc.scale(transformed_points, [angle_common_min,mach_common_min], [angle_common_max,mach_common_max])
            for i in range(number_of_values):
                #Angle of attack , Mach infinit
                mu_train.append([values[i,0], values[i,1]])
                mu_train_not_scaled.append([transformed_points[i,0], transformed_points[i,1]])
    np.save('mu_train', mu_train)
    np.save('mu_train_not_scaled', mu_train_not_scaled)
    if number_test_values > 0 and update_mu_test:
        sample = sampler.random(number_test_values)
        values = qmc.scale(sample, [angle[0],mach[0]], [angle[1],mach[1]])
        for i in range(number_test_values):
            #Angle of attack , Mach infinit
            mu_test.append([values[i,0], values[i,1]])
            mu_test_not_scaled.append([sample[i,0], sample[i,1]])
        np.save('mu_test', mu_test)
        np.save('mu_test_not_scaled', mu_test_not_scaled)
    if os.path.exists("mu_test.npy") and not update_mu_test:
        mu_test = np.load('mu_test.npy')
        mu_test_not_scaled = np.load('mu_test_not_scaled.npy')
        mu_test =  [mu.tolist() for mu in mu_test]
        mu_test_not_scaled =  [mu.tolist() for mu in mu_test_not_scaled]
    if len(mu_validation)>0:
        mu_validation_not_scaled = get_not_scale_parameters(mu_validation, angle, mach)
        np.save('mu_validation', mu_validation)
        np.save('mu_validation_not_scaled', mu_validation_not_scaled)
    plot_mu_values(mu_train, mu_test, mu_validation, 'MuValues')    
    return mu_train, mu_test, mu_train_not_scaled, mu_test_not_scaled, mu_validation_not_scaled

def get_not_scale_parameters(mu=[None], angle=[], mach=[]):
    mu_not_scaled = []
    if len(mu) > 0:
        sample = np.array(mu)
        values = qmc.scale(sample, [angle[0],mach[0]], [angle[1],mach[1]], reverse=True)
        for i in range(len(mu)):
            #Angle of attack , Mach infinit
            mu_not_scaled.append([values[i,0], values[i,1]])    
    return mu_not_scaled

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
#
# plot parameters
#
def plot_mu_values(mu_train, mu_test, mu_validation, name):
    fig, ax = plt.subplots()
    fig.set_figwidth(10.0)
    fig.set_figheight(6.0)
    regions = [
        ((min(np.array(mu_train)[:,1]), min(np.array(mu_train)[:,0])), (max(np.array(mu_train)[:,1]), max(np.array(mu_train)[:,0])), 'orange')    
        # ((0.70, 0.00), (0.71, 1.00), 'red'),    
        # ((0.71, 0.00), (0.73, 0.50), 'red'),    
        # ((0.73, 0.00), (0.74, 1.00), 'green'),  
        # ((0.74, 0.00), (0.75, 1.00), 'yellow'),  
        # ((0.75, 0.00), (0.76, 1.00), 'magenta'),  
        # ((0.71, 0.50), (0.73, 1.00), 'blue'),   
        # ((0.70, 1.00), (0.73, 1.75), 'purple'),  
        # ((0.73, 1.00), (0.74, 1.75), 'orange'),  
        # ((0.74, 1.00), (0.75, 1.75), 'brown'),  
        # ((0.75, 1.00), (0.76, 1.75), 'pink'),  
        # ((0.70, 1.75), (0.73, 2.50), 'cyan'),  
        # ((0.73, 1.75), (0.74, 2.50), 'gray'),  
        # ((0.74, 1.75), (0.75, 2.50), 'lime'),  
        # ((0.75, 1.75), (0.76, 2.50), 'olive')  
    ]
    for bottom_left, top_right, color in regions:
        rect = plt.Rectangle(bottom_left, top_right[0] - bottom_left[0], top_right[1] - bottom_left[1], 
                            facecolor=color, edgecolor=color, alpha=0.25)
        ax.add_patch(rect)
    if len(mu_train) > 0: ax.plot(np.array(mu_train)[:,1], np.array(mu_train)[:,0], 'bs', label="Train Values", markersize=4)
    if len(mu_test ) > 0: ax.plot(np.array(mu_test)[:,1], np.array(mu_test)[:,0], 'rx', label="Test Values", markersize=7)
    if len(mu_validation ) > 0: ax.plot(np.array(mu_validation)[:,1], np.array(mu_validation)[:,0], 'g*', label="Validation Values", markersize=10)
    # ax.set_xlim(min(np.array(mu_train)[:,1])-0.0025, max(np.array(mu_train)[:,1])+0.0025)
    # ax.set_ylim(min(np.array(mu_train)[:,0])-0.05, max(np.array(mu_train)[:,0])+0.05)
    plt.title('Mu Values')
    plt.ylabel('Alpha')
    plt.xlabel('Mach')
    plt.grid(True)
    plt.legend(bbox_to_anchor=(1.02, 1), loc='upper left', borderaxespad=0.)
    plt.tight_layout()
    plt.savefig(f"{name}.pdf", bbox_inches='tight', dpi=400)
    plt.close('all')

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
#
# load parameters
#
def load_mu_parameters(name):
    filename = f'mu_{name}.npy'
    if os.path.exists(filename):
        mu_npy = np.load(filename)
        mu =  [mu.tolist() for mu in mu_npy]
    else:
        mu = []
    return mu

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

def CustomizeSimulation(cls, global_model, parameters, mu):

    class CustomSimulation(cls):

        def __init__(self, model,project_parameters):
            if self._GetSimulationName() == "::[ROM Simulation]:: ": # Global ROM
                parameters["solver_settings"]["convergence_criterion"].SetString("solution_criterion")
                parameters["solver_settings"]["solving_strategy_settings"]["type"].SetString("newton_raphson")
            super().__init__(model,project_parameters)
        
        def Initialize(self):
            super().Initialize()
            if self._GetSimulationName() == "Analysis": # FOM
                    for elem in self.model["MainModelPart"].Elements:
                        elem.SetValue(CPFApp.SAVE_UPWIND_ELEMENT, True)

        def Run(self):
            angle = parameters["processes"]["boundary_conditions_process_list"][0]["Parameters"]["angle_of_attack"].GetDouble()
            mach  = parameters["processes"]["boundary_conditions_process_list"][0]["Parameters"]["mach_infinity"].GetDouble()
            parameters_number = parameters["solver_settings"]["solving_strategy_settings"]["advanced_settings"]["name"].GetString()
            parameters["solver_settings"]["solving_strategy_settings"]["advanced_settings"]["name"].SetString("")

            case_name = f'{angle}, {mach}'
            info_steps_list = []
            error = 0
            modes = 0

            if self._GetSimulationName() == "Analysis": # FOM
                start_time = time.time()
                self.Initialize()
                self.RunSolutionLoop()
                self.Finalize()

                exe_time = time.time() - start_time

                if parameters["output_processes"].Has("gid_output"):
                    simulation_name = parameters["output_processes"]["gid_output"][0]["Parameters"]["output_name"].GetString().removeprefix('Results/')

                    for process in self._GetListOfOutputProcesses():
                            if isinstance(process, CalculateRomBasisOutputProcess):
                                BasisOutputProcess = process

                    if 'Fit' in simulation_name:
                        case_type = 'train_fom'
                    elif 'Test' in simulation_name:
                        case_type = 'test_fom'
                        if os.path.exists('rom_data/RightBasisMatrix.npy'): 
                            modes = np.load('rom_data/RightBasisMatrix.npy').shape[1]
                    elif 'Run' in simulation_name:
                        case_type = 'run_fom' 
                    skin_data_filename = f"FOM_Skin_Data/{case_name}.dat"
                    fom = BasisOutputProcess._GetSnapshotsMatrix()
                    np.save(f'FOM_Snapshots/{case_name}',fom)

                    fout = open(skin_data_filename,'w')
                    modelpart = self.model["MainModelPart.Body2D_Body"]
                    for node in modelpart.Nodes:
                        x = node.X ; y = node.Y ; z = node.Z
                        cp = node.GetValue(KratosMultiphysics.PRESSURE_COEFFICIENT)
                        fout.write("%s %s %s %s\n" %(x,y,z,cp))
                    fout.close()

                    modelpart = self.model["MainModelPart"]
                    fout = open("trailing_edge_element_id.txt",'a')
                    for elem in modelpart.Elements:
                        if elem.GetValue(CPFApp.TRAILING_EDGE):
                            fout.write("%s\n" %(elem.Id))
                        if elem.GetValue(CPFApp.KUTTA):
                            fout.write("%s\n" %(elem.Id))
                        if elem.GetValue(CPFApp.WAKE):
                            fout.write("%s\n" %(elem.Id))
                    fout.close()

                    fout = open("upwind_elements_list.txt",'a')
                    for elem in modelpart.Elements:
                        if elem.GetValue(CPFApp.ID_UPWIND_ELEMENT) != 0.0:
                            fout.write("%s\n" %(elem.Id))
                            fout.write("%s\n" %(elem.GetValue(CPFApp.ID_UPWIND_ELEMENT)))
                    fout.close()
                
                    info_steps_list.append([parameters_number,
                                            case_type,
                                            angle,
                                            mach, 
                                            self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.NL_ITERATION_NUMBER],
                                            self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.RESIDUAL_NORM],
                                            error,
                                            modes,
                                            round(exe_time, 2),
                                            self.model["MainModelPart"].NumberOfNodes(),
                                            self.model["MainModelPart"].ProcessInfo[CPFApp.LIFT_COEFFICIENT],
                                            self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.DRAG_COEFFICIENT]])
                    
                    if os.path.exists(f'FOM_data.xlsx'):
                        wb = openpyxl.load_workbook(f'FOM_data.xlsx')
                        hoja = wb.active
                        for item in info_steps_list:
                            hoja.append(item)
                        wb.save(f'FOM_data.xlsx')
                    else:
                        wb = openpyxl.Workbook()
                        hoja = wb.active
                        hoja.append(('P_Number', 'Case name', 'Angle [º]', 'Mach', 'NL iterations', 'Residual norm', 'Approximation error [%]', 'Modes', 'Time [sec]', 'Nodes', 'Cl', 'Cd'))
                        for item in info_steps_list:
                            hoja.append(item)
                        wb.save(f'FOM_data.xlsx')

            elif self._GetSimulationName() == "::[ROM Simulation]:: ": # Global ROM

                start_time = time.time()
                self.Initialize()
                self.RunSolutionLoop()
                self.Finalize()                
                exe_time = time.time() - start_time

                if parameters["output_processes"].Has("gid_output"):

                    simulation_name = parameters["output_processes"]["gid_output"][0]["Parameters"]["output_name"].GetString().removeprefix('Results/')

                    for process in self._GetListOfOutputProcesses():
                            if isinstance(process, CalculateRomBasisOutputProcess):
                                BasisOutputProcess = process

                    if 'HROM' in simulation_name:                      # HROM
                        if 'Fit' in simulation_name:
                            case_type = 'train_'
                        elif 'Test' in simulation_name:
                            case_type = 'test_'
                        elif 'Run' in simulation_name:
                            case_type = 'run_' 
                            
                        hrom = BasisOutputProcess._GetSnapshotsMatrix()
                        fom = np.load(f'FOM_Snapshots/{case_name}.npy')

                        if (len(fom) != len(hrom) or 'HHROM' in simulation_name):
                            q_matrix = []
                            main_model_part = self.model["MainModelPart"]
                            q_matrix.append(np.array(main_model_part.GetValue(KratosMultiphysics.RomApplication.ROM_SOLUTION_INCREMENT)).reshape(-1,1))
                            q_matrix = np.block(q_matrix)
                            phi = np.load(f'rom_data/RightBasisMatrix.npy')
                            if (q_matrix.shape[0] == 1): q_matrix = q_matrix.T
                            hrom = phi @ q_matrix
                            np.save(f'HHROM_Snapshots/{case_name}', hrom)
                            case_type = case_type + 'hhrom'
                        else:
                            np.save(f'HROM_Snapshots/{case_name}', hrom)
                            case_type = case_type + 'hrom'
                            skin_data_filename = f"HROM_Skin_Data/{case_name}.dat"
                            fout = open(skin_data_filename,'w')
                            modelpart = self.model["MainModelPart.Body2D_Body"]
                            for node in modelpart.Nodes:
                                x = node.X ; y = node.Y ; z = node.Z
                                cp = node.GetValue(KratosMultiphysics.PRESSURE_COEFFICIENT)
                                fout.write("%s %s %s %s\n" %(x,y,z,cp))
                            fout.close()

                        error = np.linalg.norm(fom-hrom)/np.linalg.norm(fom)
                        modes = np.load('rom_data/RightBasisMatrix.npy').shape[1]
                    
                        info_steps_list.append([parameters_number,
                                                case_type,
                                                angle,
                                                mach, 
                                                self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.NL_ITERATION_NUMBER],
                                                self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.RESIDUAL_NORM],
                                                error,
                                                modes,
                                                round(exe_time, 2),
                                                self.model["MainModelPart"].NumberOfNodes(),
                                                self.model["MainModelPart"].ProcessInfo[CPFApp.LIFT_COEFFICIENT],
                                                self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.DRAG_COEFFICIENT]])
                        
                        if os.path.exists(f'case_data.xlsx'):
                            wb = openpyxl.load_workbook(f'case_data.xlsx')
                            hoja = wb.active
                            for item in info_steps_list:
                                hoja.append(item)
                            wb.save(f'case_data.xlsx')
                        else:
                            wb = openpyxl.Workbook()
                            hoja = wb.active
                            hoja.append(('P_Number', 'Case name', 'Angle [º]', 'Mach', 'NL iterations', 'Residual norm', 'Approximation error [%]', 'Modes', 'Time [sec]', 'Nodes', 'Cl', 'Cd'))
                            for item in info_steps_list:
                                hoja.append(item)
                            wb.save(f'case_data.xlsx')

                    elif 'ROM' in simulation_name:                    # ROM
                        if 'Fit' in simulation_name:
                            case_type = 'train_rom'
                        elif 'Test' in simulation_name:
                            case_type = 'test_rom' 
                        elif 'Run' in simulation_name:
                            case_type = 'run_rom' 
                        modes = np.load('rom_data/RightBasisMatrix.npy').shape[1]
                        skin_data_filename = f"ROM_Skin_Data/{case_name}.dat"

                        rom = BasisOutputProcess._GetSnapshotsMatrix()
                        np.save(f'ROM_Snapshots/{case_name}',rom)
                        fom = np.load(f'FOM_Snapshots/{case_name}.npy')
                        error = np.linalg.norm(fom-rom)/np.linalg.norm(fom)

                        fout = open(skin_data_filename,'w')
                        modelpart = self.model["MainModelPart.Body2D_Body"]
                        for node in modelpart.Nodes:
                            x = node.X ; y = node.Y ; z = node.Z
                            cp = node.GetValue(KratosMultiphysics.PRESSURE_COEFFICIENT)
                            fout.write("%s %s %s %s\n" %(x,y,z,cp))
                        fout.close()
                    
                        info_steps_list.append([parameters_number,
                                                case_type,
                                                angle,
                                                mach, 
                                                self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.NL_ITERATION_NUMBER],
                                                self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.RESIDUAL_NORM],
                                                error,
                                                modes,
                                                round(exe_time, 2),
                                                self.model["MainModelPart"].NumberOfNodes(),
                                                self.model["MainModelPart"].ProcessInfo[CPFApp.LIFT_COEFFICIENT],
                                                self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.DRAG_COEFFICIENT]])
                        
                        if os.path.exists(f'case_data.xlsx'):
                            wb = openpyxl.load_workbook(f'case_data.xlsx')
                            hoja = wb.active
                            for item in info_steps_list:
                                hoja.append(item)
                            wb.save(f'case_data.xlsx')
                        else:
                            wb = openpyxl.Workbook()
                            hoja = wb.active
                            hoja.append(('P_Number', 'Case name','Angle [º]', 'Mach', 'NL iterations', 'Residual norm', 'Approximation error [%]', 'Modes', 'Time [sec]', 'Nodes', 'Cl', 'Cd'))
                            for item in info_steps_list:
                                hoja.append(item)
                            wb.save(f'case_data.xlsx')

    return CustomSimulation(global_model, parameters)


# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

def UpdateProjectParameters(parameters, mu=None):
    angle_of_attack = mu[0]
    mach_infinity   = mu[1]
    parameters["processes"]["boundary_conditions_process_list"][0]["Parameters"]["angle_of_attack"].SetDouble(np.double(angle_of_attack))
    parameters["processes"]["boundary_conditions_process_list"][0]["Parameters"]["mach_infinity"].SetDouble(np.double(mach_infinity))

    if (mach_infinity > 0.73 and mach_infinity <= 0.74 and angle_of_attack >= 0.0 and angle_of_attack <= 1.00):
        # input("2")
        parameters["solver_settings"]["solving_strategy_settings"]["advanced_settings"]["name"].SetString('2')
        parameters["solver_settings"]["scheme_settings"]["initial_critical_mach"].SetDouble(0.92)
        parameters["solver_settings"]["scheme_settings"]["initial_upwind_factor_constant"].SetDouble(1.8)
        parameters["solver_settings"]["scheme_settings"]["update_relative_residual_norm"].SetDouble(1e-3)
        parameters["solver_settings"]["scheme_settings"]["target_critical_mach"].SetDouble(0.95)
        parameters["solver_settings"]["scheme_settings"]["target_upwind_factor_constant"].SetDouble(1.3)

    if (mach_infinity > 0.74 and mach_infinity <= 0.75 and angle_of_attack >= 0.0 and angle_of_attack <= 1.00):
        # input("3")
        parameters["solver_settings"]["solving_strategy_settings"]["advanced_settings"]["name"].SetString('3')
        parameters["solver_settings"]["scheme_settings"]["initial_critical_mach"].SetDouble(0.90)
        parameters["solver_settings"]["scheme_settings"]["initial_upwind_factor_constant"].SetDouble(2.0)
        parameters["solver_settings"]["scheme_settings"]["update_relative_residual_norm"].SetDouble(1e-3)
        parameters["solver_settings"]["scheme_settings"]["target_critical_mach"].SetDouble(0.93)
        parameters["solver_settings"]["scheme_settings"]["target_upwind_factor_constant"].SetDouble(1.5)

    if (mach_infinity > 0.75 and mach_infinity <= 0.76 and angle_of_attack >= 0.0 and angle_of_attack <= 1.00):
        # input("4")
        parameters["solver_settings"]["solving_strategy_settings"]["advanced_settings"]["name"].SetString('4')
        parameters["solver_settings"]["scheme_settings"]["initial_critical_mach"].SetDouble(0.90)
        parameters["solver_settings"]["scheme_settings"]["initial_upwind_factor_constant"].SetDouble(2.5)
        parameters["solver_settings"]["scheme_settings"]["update_relative_residual_norm"].SetDouble(1e-3)
        parameters["solver_settings"]["scheme_settings"]["target_critical_mach"].SetDouble(0.95)
        parameters["solver_settings"]["scheme_settings"]["target_upwind_factor_constant"].SetDouble(2.0)

    if (mach_infinity > 0.71 and mach_infinity <= 0.73 and angle_of_attack > 0.5 and angle_of_attack <= 1.00):
        # input("5")
        parameters["solver_settings"]["solving_strategy_settings"]["advanced_settings"]["name"].SetString('5')
        parameters["solver_settings"]["scheme_settings"]["initial_critical_mach"].SetDouble(0.90)
        parameters["solver_settings"]["scheme_settings"]["initial_upwind_factor_constant"].SetDouble(1.8)
        parameters["solver_settings"]["scheme_settings"]["update_relative_residual_norm"].SetDouble(1e-3)
        parameters["solver_settings"]["scheme_settings"]["target_critical_mach"].SetDouble(0.95)
        parameters["solver_settings"]["scheme_settings"]["target_upwind_factor_constant"].SetDouble(1.5)

    if (mach_infinity >= 0.70 and mach_infinity <= 0.73 and angle_of_attack > 1.0 and angle_of_attack <= 1.75):
        # input("6")
        parameters["solver_settings"]["solving_strategy_settings"]["advanced_settings"]["name"].SetString('6')
        parameters["solver_settings"]["scheme_settings"]["initial_critical_mach"].SetDouble(0.90)
        parameters["solver_settings"]["scheme_settings"]["initial_upwind_factor_constant"].SetDouble(2.5)
        parameters["solver_settings"]["scheme_settings"]["update_relative_residual_norm"].SetDouble(1e-3)
        parameters["solver_settings"]["scheme_settings"]["target_critical_mach"].SetDouble(0.95)
        parameters["solver_settings"]["scheme_settings"]["target_upwind_factor_constant"].SetDouble(2.0)

    if (mach_infinity > 0.73 and mach_infinity <= 0.74 and angle_of_attack > 1.0 and angle_of_attack <= 1.75):
        # input("7")
        parameters["solver_settings"]["solving_strategy_settings"]["advanced_settings"]["name"].SetString('7')
        parameters["solver_settings"]["scheme_settings"]["initial_critical_mach"].SetDouble(0.90)
        parameters["solver_settings"]["scheme_settings"]["initial_upwind_factor_constant"].SetDouble(3.0)
        parameters["solver_settings"]["scheme_settings"]["update_relative_residual_norm"].SetDouble(1e-3)
        parameters["solver_settings"]["scheme_settings"]["target_critical_mach"].SetDouble(0.95)
        parameters["solver_settings"]["scheme_settings"]["target_upwind_factor_constant"].SetDouble(2.0)

    if (mach_infinity > 0.74 and mach_infinity <= 0.75 and angle_of_attack > 1.0 and angle_of_attack <= 1.75):
        # input("8")
        parameters["solver_settings"]["solving_strategy_settings"]["advanced_settings"]["name"].SetString('8')
        parameters["solver_settings"]["scheme_settings"]["initial_critical_mach"].SetDouble(0.85)
        parameters["solver_settings"]["scheme_settings"]["initial_upwind_factor_constant"].SetDouble(3.0)
        parameters["solver_settings"]["scheme_settings"]["update_relative_residual_norm"].SetDouble(1e-3)
        parameters["solver_settings"]["scheme_settings"]["target_critical_mach"].SetDouble(0.95)
        parameters["solver_settings"]["scheme_settings"]["target_upwind_factor_constant"].SetDouble(2.0)

    if (mach_infinity > 0.75 and mach_infinity <= 0.76 and angle_of_attack > 1.0 and angle_of_attack <= 1.75):
        # input("9")
        parameters["solver_settings"]["solving_strategy_settings"]["advanced_settings"]["name"].SetString('9')
        parameters["solver_settings"]["scheme_settings"]["initial_critical_mach"].SetDouble(0.80)
        parameters["solver_settings"]["scheme_settings"]["initial_upwind_factor_constant"].SetDouble(3.5)
        parameters["solver_settings"]["scheme_settings"]["update_relative_residual_norm"].SetDouble(1e-3)
        parameters["solver_settings"]["scheme_settings"]["target_critical_mach"].SetDouble(0.86)
        parameters["solver_settings"]["scheme_settings"]["target_upwind_factor_constant"].SetDouble(3.0)

    if (mach_infinity >= 0.70 and mach_infinity <= 0.73 and angle_of_attack > 1.75 and angle_of_attack <= 2.50):
        # input("10)
        parameters["solver_settings"]["solving_strategy_settings"]["advanced_settings"]["name"].SetString('10')
        parameters["solver_settings"]["scheme_settings"]["initial_critical_mach"].SetDouble(0.85)
        parameters["solver_settings"]["scheme_settings"]["initial_upwind_factor_constant"].SetDouble(4.0)
        parameters["solver_settings"]["scheme_settings"]["update_relative_residual_norm"].SetDouble(1e-3)
        parameters["solver_settings"]["scheme_settings"]["target_critical_mach"].SetDouble(0.90)
        parameters["solver_settings"]["scheme_settings"]["target_upwind_factor_constant"].SetDouble(2.0)


    if (mach_infinity > 0.73 and mach_infinity <= 0.74 and angle_of_attack > 1.75 and angle_of_attack <= 2.50):
        # input("11")
        parameters["solver_settings"]["solving_strategy_settings"]["advanced_settings"]["name"].SetString('11')
        parameters["solver_settings"]["scheme_settings"]["initial_critical_mach"].SetDouble(0.80)
        parameters["solver_settings"]["scheme_settings"]["initial_upwind_factor_constant"].SetDouble(4.0)
        parameters["solver_settings"]["scheme_settings"]["update_relative_residual_norm"].SetDouble(1e-3)
        parameters["solver_settings"]["scheme_settings"]["target_critical_mach"].SetDouble(0.90)
        parameters["solver_settings"]["scheme_settings"]["target_upwind_factor_constant"].SetDouble(2.0)

    if (mach_infinity > 0.74 and mach_infinity <= 0.75 and angle_of_attack > 1.75 and angle_of_attack <= 2.50):
        # input("12")
        parameters["solver_settings"]["solving_strategy_settings"]["advanced_settings"]["name"].SetString('12')
        parameters["solver_settings"]["scheme_settings"]["initial_critical_mach"].SetDouble(0.80)
        parameters["solver_settings"]["scheme_settings"]["initial_upwind_factor_constant"].SetDouble(5.0)
        parameters["solver_settings"]["scheme_settings"]["update_relative_residual_norm"].SetDouble(1e-3)
        parameters["solver_settings"]["scheme_settings"]["target_critical_mach"].SetDouble(0.85)
        parameters["solver_settings"]["scheme_settings"]["target_upwind_factor_constant"].SetDouble(3.0)

    if (mach_infinity > 0.75 and mach_infinity <= 0.76 and angle_of_attack > 1.75 and angle_of_attack <= 2.50):
        # input("13")
        parameters["solver_settings"]["solving_strategy_settings"]["advanced_settings"]["name"].SetString('13')
        parameters["solver_settings"]["scheme_settings"]["initial_critical_mach"].SetDouble(0.80)
        parameters["solver_settings"]["scheme_settings"]["initial_upwind_factor_constant"].SetDouble(7.0)
        parameters["solver_settings"]["scheme_settings"]["update_relative_residual_norm"].SetDouble(1e-3)
        parameters["solver_settings"]["scheme_settings"]["target_critical_mach"].SetDouble(0.85)
        parameters["solver_settings"]["scheme_settings"]["target_upwind_factor_constant"].SetDouble(5.5)

    return parameters

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

def UpdateMaterialParametersFile(material_parametrs_file_name, mu):
    pass

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #



def GetRomManagerParameters():
    general_rom_manager_parameters = KratosMultiphysics.Parameters("""{
            "rom_stages_to_train" : ["ROM","HHROM"],            // ["FOM","ROM","HROM","HHROM"]
            "rom_stages_to_test"  : ["ROM","HHROM"],            // ["FOM","ROM","HROM","HHROM"]
            "paralellism" : null,                       // null, TODO: add "compss"
            "projection_strategy": "galerkin",          // "lspg", "galerkin", "petrov_galerkin"
            "type_of_decoder" : "linear",               // "linear" "ann_enhanced",  TODO: add "quadratic"
            "assembling_strategy": "global",            // "global", "elemental"
            "save_gid_output": true,                    // false, true #if true, it must exits previously in the ProjectParameters.json
            "save_vtk_output": true,                   // false, true #if true, it must exits previously in the ProjectParameters.json
            "output_name": "id",                        // "id" , "mu"
            "store_nonconverged_fom_solutions": true,
            "ROM":{
                "analysis_stage" : "KratosMultiphysics.CompressiblePotentialFlowApplication.potential_flow_analysis",
                "svd_truncation_tolerance": 1e-12,
                "print_singular_values": true,
                "use_non_converged_sols" : false,
                "model_part_name": "MainModelPart",                         // This changes depending on the simulation: Structure, FluidModelPart, ThermalPart #TODO: Idenfity it automatically
                "nodal_unknowns": ["VELOCITY_POTENTIAL", "AUXILIARY_VELOCITY_POTENTIAL"],     // Main unknowns. Snapshots are taken from these
                "rom_basis_output_format": "numpy",
                "rom_basis_output_name": "RomParameters",
                "rom_basis_output_folder": "rom_data",
                "snapshots_control_type": "step",                          // "step", "time"
                "snapshots_interval": 1,
                "galerkin_rom_bns_settings": {
                    "monotonicity_preserving": false
                },
                "lspg_rom_bns_settings": {
                    "train_petrov_galerkin": false,
                    "basis_strategy": "reactions",                        // 'residuals', 'jacobian', 'reactions'
                    "include_phi": false,
                    "svd_truncation_tolerance": 0,
                    "solving_technique": "normal_equations",              // 'normal_equations', 'qr_decomposition'
                    "monotonicity_preserving": false
                },
                "petrov_galerkin_rom_bns_settings": {
                    "monotonicity_preserving": false
                }
            },
            "HROM":{
                "element_selection_type": "empirical_cubature",
                "element_selection_svd_truncation_tolerance": 0,
                "create_hrom_visualization_model_part" : true,
                "echo_level" : 1,                                       
                "hrom_format": "numpy",
                "initial_candidate_elements_model_part_list" : [],
                "initial_candidate_conditions_model_part_list" : [],
                "include_nodal_neighbouring_elements_model_parts_list":[],
                "include_elements_model_parts_list": [],
                "include_conditions_model_parts_list": ["MainModelPart.PotentialWallCondition2D_Far_field_Auto1","MainModelPart.Body2D_Body"],
                "include_minimum_condition": false,
                "include_condition_parents": true
            }
        }""")

    return general_rom_manager_parameters


# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #


if __name__ == "__main__":

    folder_names = ["FOM_Snapshots"  , "FOM_Skin_Data", 
                    "ROM_Snapshots"  , "ROM_Skin_Data", 
                    "HROM_Snapshots" , "HROM_Skin_Data",
                    "HHROM_Snapshots", "HHROM_Skin_Data", 
                    "RBF_Snapshots"  , "RBF_Skin_Data", 
                    "Train_Captures" , "Test_Captures", "Validation"]
    for name in folder_names:
        if not os.path.exists(name):
            os.mkdir(name)

    ###############################
    # PARAMETERS SETTINGS
    update_parameters  = False
    update_mu_test     = True
    VALIDATION         = True
    number_of_mu_test  = 15
    alpha              = 1.0
    beta               = 1.0
    update_residuals   = False
    mach_range         = [0.71, 0.74]
    angle_range        = [0.50, 1.50]
    ################################

    regions = [
        ((0.71, 0.73),(0.00, 0.50), 10),#, 1  'red'    
        ((0.70, 0.71),(0.00, 1.00), 10),#, 1  'red'    
        ((0.73, 0.74),(0.00, 1.00), 20),#, 2  'green'  
        ((0.74, 0.75),(0.00, 1.00), 15),#, 3  'yellow' 
        ((0.75, 0.76),(0.00, 1.00), 15),#, 4  'magenta'
        ((0.71, 0.73),(0.50, 1.00), 30),#, 5  'blue'   
        ((0.70, 0.73),(1.00, 1.75), 30),#, 6  'purple'  
        ((0.73, 0.74),(1.00, 1.75), 20),#, 7  'orange'  
        ((0.74, 0.75),(1.00, 1.75), 15),#, 8  'brown'  
        ((0.75, 0.76),(1.00, 1.75), 15),#, 9  'pink'  
        ((0.70, 0.73),(1.75, 2.50), 25),#, 10 'cyan'  
        ((0.73, 0.74),(1.75, 2.50), 15),#, 11 'gray'  
        ((0.74, 0.75),(1.75, 2.50), 15),#, 12 'lime'  
        ((0.75, 0.76),(1.75, 2.50), 15) #  13 'olive'  
    ]

    mu_validation = []
    mu_validation_not_scaled = []
    if VALIDATION:
        if angle_range[0] <= 1.0 and angle_range[1]>= 1.0 and mach_range[0] <= 0.72 and mach_range[1]>=0.72:
            mu_validation.append([1.0,0.72])
        if angle_range[0] <= 1.0 and angle_range[1]>= 1.0 and mach_range[0] <= 0.73 and mach_range[1]>=0.73:
            mu_validation.append([1.0,0.73])
        if angle_range[0] <= 1.0 and angle_range[1]>= 1.0 and mach_range[0] <= 0.75 and mach_range[1]>=0.75:
            mu_validation.append([1.0,0.75])
        if angle_range[0] <= 2.0 and angle_range[1]>= 2.0 and mach_range[0] <= 0.75 and mach_range[1]>=0.75:
            mu_validation.append([2.0,0.75])
    
    if update_parameters:
        (mu_train, mu_test,
        mu_train_not_scaled, mu_test_not_scaled,
        mu_validation_not_scaled) = get_multiple_parameters(regions             = regions           ,
                                                            number_test_values  = number_of_mu_test , 
                                                            angle               = angle_range       , 
                                                            mach                = mach_range        , 
                                                            mu_validation       = mu_validation     ,
                                                            method              = 'Halton'          ,
                                                            update_mu_test      = update_mu_test    ,
                                                            alpha = alpha, beta = beta )
        KratosMultiphysics.kratos_utilities.DeleteFileIfExisting('upwind_elements_list.txt')
        KratosMultiphysics.kratos_utilities.DeleteFileIfExisting('trailing_edge_element_id.txt')
        KratosMultiphysics.kratos_utilities.DeleteFileIfExisting('case_data.xlsx')
    else:
        if update_residuals:
            KratosMultiphysics.kratos_utilities.DeleteDirectoryIfExisting('ResidualsSnapshotsMatrix.zarr')
        KratosMultiphysics.kratos_utilities.DeleteFileIfExisting('case_data.xlsx')
        KratosMultiphysics.kratos_utilities.DeleteDirectoryIfExisting('Train_Captures')
        KratosMultiphysics.kratos_utilities.DeleteDirectoryIfExisting('Test_Captures')
        KratosMultiphysics.kratos_utilities.DeleteDirectoryIfExisting('Validation')
        for name in folder_names:
            if not os.path.exists(name):
                os.mkdir(name)
        mu_train_list = load_mu_parameters('train')
        mu_train_not_scaled_list = load_mu_parameters('train_not_scaled')
        mu_test_list = load_mu_parameters('test')
        mu_test_not_scaled_list = load_mu_parameters('test_not_scaled')
        mu_validation_list = load_mu_parameters('validation')
        mu_validation_not_scaled_list = load_mu_parameters('validation_not_scaled')
        
        mu_train = []; mu_test = []; mu_validation = []
        for angle, mach in mu_train_list:
            if (    angle <= angle_range[1]
                and angle >= angle_range[0] 
                and mach  <=  mach_range[1]
                and mach  >=  mach_range[0]):
                mu_train.append([angle, mach])
        mu_train_not_scaled = [mu_value for mu_value, mu_train_value in zip(mu_train_not_scaled_list, mu_train) if mu_train_value in mu_train_list]

        for angle, mach in mu_test_list:
            if (    angle <= angle_range[1]
                and angle >= angle_range[0] 
                and mach  <=  mach_range[1]
                and mach  >=  mach_range[0]):
                mu_test.append([angle, mach])
        mu_test_not_scaled = [mu_value for mu_value, mu_test_value in zip(mu_test_not_scaled_list, mu_test) if mu_test_value in mu_test_list]

        for angle, mach in mu_validation_list:
            if (    angle <= angle_range[1]
                and angle >= angle_range[0] 
                and mach  <=  mach_range[1]
                and mach  >=  mach_range[0]):
                mu_validation.append([angle, mach])
        mu_validation_not_scaled = [mu_value for mu_value, mu_validation_value in zip(mu_validation_not_scaled_list, mu_validation) if mu_validation_value in mu_validation_list]

        plot_mu_values(mu_train, mu_test, mu_validation, 'MuValues')

    print('Number of train cases: ', len(mu_train))
    input('Pause')

    general_rom_manager_parameters = GetRomManagerParameters()
    project_parameters_name = "ProjectParameters.json"

    rom_manager = RomManager(project_parameters_name,general_rom_manager_parameters,
                             CustomizeSimulation,UpdateProjectParameters,UpdateMaterialParametersFile,
                             relaunch_FOM=False, relaunch_ROM=True, relaunch_HROM=True, rebuild_phi=True)

    rom_manager.Fit(mu_train)

    rom_manager.Test(mu_test)

    if VALIDATION:
        for mu in mu_validation:
            if not os.path.exists(f'FOM_Snapshots/{mu[0]},{mu[1]}.npy'):
                rom_manager.RunFOM(mu_run=[mu])
        training_stages = rom_manager.general_rom_manager_parameters["rom_stages_to_test"].GetStringArray()
        if any(item == "ROM" for item in training_stages):
            rom_manager.RunROM(mu_run=mu_validation)
        if any(item == "HROM" for item in training_stages):
            rom_manager.RunHROM(mu_run=mu_validation,use_full_model_part=True)
        if any(item == "HHROM" for item in training_stages):
            rom_manager.RunHHROM(mu_run=mu_validation)

    rom_manager.PrintErrors(show_q_errors=False)