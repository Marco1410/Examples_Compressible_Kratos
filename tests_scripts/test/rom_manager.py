import os
import time
import openpyxl
import tracemalloc
import numpy as np
from parameters_manager import * 
import KratosMultiphysics
import KratosMultiphysics.kratos_utilities
import KratosMultiphysics.CompressiblePotentialFlowApplication as CPFApp
from KratosMultiphysics.RomApplication.rom_manager import RomManager
from KratosMultiphysics.RomApplication.rom_database import RomDatabase
from KratosMultiphysics.RomApplication.calculate_rom_basis_output_process import CalculateRomBasisOutputProcess



# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

def CustomizeSimulation(cls, global_model, parameters, mu, constant):

    class CustomSimulation(cls):

        def __init__(self, model,project_parameters):
            super().__init__(model,project_parameters)
            if len(mu) > 0:
                self.angle               = mu[0]
                self.mach                = mu[1]
                self.weight_constant     = mu[2]
                self.density_constant    = mu[3]
                self.mu_id               = mu[4]
                self.svd_tolerance       = mu[5]
                self.rom_solver_strategy = mu[6]
                if self.rom_solver_strategy: 
                    self.strategy = 'newton_raphson'
                else:
                    self.strategy = 'line_search'
                self.case_name           = f'{self.angle}, {self.mach}'
                self.simulation_name     = []

            if parameters["output_processes"].Has("gid_output"):
                self.simulation_name = parameters["output_processes"]["gid_output"][0]["Parameters"]["output_name"].GetString().removeprefix('Results/')

                if 'ROM' in self.simulation_name: # Use Line Search or Newton Raphson for rom simulations
                    parameters["solver_settings"]["solving_strategy_settings"]["type"].SetString(self.strategy)
            
        def GetFinalData(self):
            Cp_final_data = []
            modelpart = self.model["MainModelPart.Body2D_Body"]
            for node in modelpart.Nodes:
                Cp_final_data.append([node.X, node.Y, node.Z, node.GetValue(KratosMultiphysics.PRESSURE_COEFFICIENT)])

            Full_Cp_final_data = []
            modelpart = self.model["MainModelPart"]
            for node in modelpart.Nodes:
                Full_Cp_final_data.append([node.X, node.Y, node.Z, node.GetValue(KratosMultiphysics.PRESSURE_COEFFICIENT)])

            return {'Cp_data': np.array(Cp_final_data),
                    'Full_Cp_data': np.array(Full_Cp_final_data),
                    'Residual_norm': self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.RESIDUAL_NORM],
                    'NL_iterations': self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.NL_ITERATION_NUMBER],
                    'Cl': self.model["MainModelPart"].ProcessInfo[CPFApp.LIFT_COEFFICIENT],
                    'Cd': self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.DRAG_COEFFICIENT],
                    'Memory': self.used_memory_mb,
                    'Time': self.exe_time}
            

        def Run(self):

            # Init memory monitoring
            tracemalloc.start()
            self.Initialize()

            for elem in self.model["MainModelPart"].Elements: #Initilize of elemental variable
                elem.SetValue(KratosMultiphysics.ACTIVATION_LEVEL, 0)
            
            if 'FOM' in self.simulation_name: #MARK ELEMENTS
                self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.REFINEMENT_LEVEL] = 1

            if 'ROM' in self.simulation_name: #USE WEIGHT VERSION
                self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.REFINEMENT_LEVEL] =  1 if self.weight_constant > 0 else 0

            # SET WEIGHT CONSTANT
            if 'ROM' in self.simulation_name:
                self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.CONSTRAINT_SCALE_FACTOR] = self.weight_constant

            # Init time control
            start_time = time.time()
            self.RunSolutionLoop()
            self.exe_time = round(time.time() - start_time, 2)

            self.Finalize()

            used_memory = tracemalloc.get_traced_memory()
            tracemalloc.stop()
            self.used_memory_mb = (used_memory[1] - used_memory[0]) / 1024 ** 2  # Convert to MB

            if parameters["output_processes"].Has("gid_output"):
                info_steps_list = []
                u_error = 0 
                modes   = 0

                for process in self._GetListOfOutputProcesses():
                    if isinstance(process, CalculateRomBasisOutputProcess):
                        BasisOutputProcess = process

                if 'FOM' in self.simulation_name:
                    case = 'FOM'
                    if 'Fit' in self.simulation_name:
                        case_type = 'train_fom'
                    elif 'Test' in self.simulation_name:
                        case_type = 'test_fom'
                    elif 'Run' in self.simulation_name:
                        case_type = 'run_fom' 

                    modelpart = self.model["MainModelPart"]
                    if not os.path.exists("Selected_elements_lists"):
                        os.mkdir("Selected_elements_lists")
                    fout = open(f"Selected_elements_lists/{self.case_name}.txt",'a') #Elements list for rom approximation
                    for elem in modelpart.Elements:
                        if elem.GetValue(KratosMultiphysics.ACTIVATION_LEVEL):
                            fout.write("%s\n" %(elem.Id))
                    fout.close()
                    
                elif 'ROM' in self.simulation_name:
                    case = 'ROM'
                    if 'Fit' in self.simulation_name:
                        case_type = 'train_rom'
                    elif 'Test' in self.simulation_name:
                        case_type = 'test_rom' 
                    elif 'Run' in self.simulation_name:
                        case_type = 'run_rom' 

                    u_rom = BasisOutputProcess._GetSnapshotsMatrix()

                    with open("rom_manager_parameters.json", 'r') as parameter_file:
                        general_rom_manager_parameters = KratosMultiphysics.Parameters(parameter_file.read())

                    general_rom_manager_parameters["ROM"]["svd_truncation_tolerance"].SetDouble(self.svd_tolerance)

                    data_base = RomDatabase(general_rom_manager_parameters, mu_names=['Alpha','Mach','Weight','Beta','MuId','SVDTolerance','RomSolverStrategy'])
                    u_fom = data_base.get_snapshots_matrix_from_database([mu], table_name=f'FOM')
                                
                    u_error = np.linalg.norm(u_fom-u_rom)/np.linalg.norm(u_fom)

                if os.path.exists('rom_data/RightBasisMatrix.npy'): 
                    modes = np.load('rom_data/RightBasisMatrix.npy').shape[1]

                info_steps_list.append([case_type,
                                        self.mu_id,
                                        self.angle,
                                        self.mach, 
                                        self.weight_constant, 
                                        self.density_constant, 
                                        self.strategy,
                                        self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.RESIDUAL_NORM],
                                        self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.NL_ITERATION_NUMBER],
                                        u_error,
                                        modes,
                                        self.svd_tolerance,
                                        self.exe_time,
                                        self.model["MainModelPart"].NumberOfNodes(),
                                        self.model["MainModelPart"].ProcessInfo[CPFApp.LIFT_COEFFICIENT],
                                        self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.DRAG_COEFFICIENT],
                                        self.used_memory_mb])
                
                if os.path.exists(f'{case}_data.xlsx'):
                    wb = openpyxl.load_workbook(f'{case}_data.xlsx')
                    hoja = wb.active
                    for item in info_steps_list:
                        hoja.append(item)
                    wb.save(f'{case}_data.xlsx')
                else:
                    wb = openpyxl.Workbook()
                    hoja = wb.active
                    hoja.append(('Case name', 'Mu id', 'Angle [º]', 'Mach', 'Weight', 'Density constant', 'Rom solver strategy', 'Residual norm', 
                                 'NL iter','u relative error', 'Modes', 'SVD tol', 'Time [sec]', 'Nodes', 'Cl', 'Cd', 'Memory [Mb]'))
                    for item in info_steps_list:
                        hoja.append(item)
                    wb.save(f'{case}_data.xlsx')

    return CustomSimulation(global_model, parameters)


# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

def UpdateProjectParameters(parameters, mu=None):
    angle_of_attack = mu[0]
    mach_infinity   = mu[1]
    parameters["processes"]["boundary_conditions_process_list"][0]["Parameters"]["angle_of_attack"].SetDouble(np.double(angle_of_attack))
    parameters["processes"]["boundary_conditions_process_list"][0]["Parameters"]["mach_infinity"].SetDouble(np.double(mach_infinity))

    return parameters

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

def UpdateMaterialParametersFile(material_parametrs_file_name, mu):
    pass

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

def GetRomManagerParameters():
    with open("rom_manager_parameters.json", 'r') as parameter_file:
        general_rom_manager_parameters = KratosMultiphysics.Parameters(parameter_file.read())

    return general_rom_manager_parameters

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

def extract_unique_elements(mu_train, folder='Selected_elements_lists', output_file='selected_elements_list.txt'):
    unique_elements = set()
    for mu in mu_train:
        filename = f"{mu[0]}, {mu[1]}.txt"
        filepath = os.path.join(folder, filename)
        if os.path.exists(filepath):
            with open(filepath, 'r') as f:
                for line in f:
                    try:
                        unique_elements.add(int(line.strip()))
                    except ValueError:
                        pass  
    with open(output_file, 'w') as f:
        for elem in sorted(unique_elements):
            f.write(f"{elem}\n")
    
# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #



if __name__ == "__main__":

    ###########################################################
    # PARAMETERS SETTINGS
    update_parameters         = True
    number_of_mu_train        = [5]
    number_of_mu_test         = 3
    mach_range                = [0.70, 0.75]
    angle_range               = [0.00, 2.00]
    ###########################################################
    alpha_values              = [1.0]
    beta_values               = [1.0]
    svd_truncation_tolerances = [1e-12]
    rom_solver_strategies     = [1] #['line_search', 'newton_raphson']
    weight_constants          = [1.0]
    ###########################################################
    relaunch_FOM              = True 
    relaunch_ROM              = True
    rebuild_phi               = True 
    VALIDATION                = True
    TEST                      = True
    project_parameters_name   = "ProjectParametersNaca0012.json"
    ###########################################################

    mu_validation = []
    if VALIDATION:
        if angle_range[0] <= 1.0 and angle_range[1] >= 1.0 and mach_range[0] <= 0.72 and mach_range[1] >= 0.72:
            mu_validation.append([1.0,0.72,1.0,1.0,1.0,1.0,1.0])
        if angle_range[0] <= 1.0 and angle_range[1] >= 1.0 and mach_range[0] <= 0.73 and mach_range[1] >= 0.73:
            mu_validation.append([1.0,0.73,1.0,1.0,1.0,1.0,1.0])
        if angle_range[0] <= 1.0 and angle_range[1] >= 1.0 and mach_range[0] <= 0.75 and mach_range[1] >= 0.75:
            mu_validation.append([1.0,0.75,1.0,1.0,1.0,1.0,1.0])
        if angle_range[0] <= 2.0 and angle_range[1] >= 2.0 and mach_range[0] <= 0.75 and mach_range[1] >= 0.75:
            mu_validation.append([2.0,0.75,1.0,1.0,1.0,1.0,1.0])

    print(':::::::::::::::::::::::::::::::::::::::::')
    print('Global simulations: ', len(number_of_mu_train)*len(beta_values)*len(svd_truncation_tolerances)*len(rom_solver_strategies)*len(weight_constants))
    print('Number of FOM simulations: ', sum(number_of_mu_train)*len(beta_values)+number_of_mu_test+len(mu_validation))
    print('Number of ROM simulations: ', len(beta_values)*(sum(number_of_mu_train)+len(number_of_mu_train)*(number_of_mu_test+len(mu_validation)))*len(svd_truncation_tolerances)*len(rom_solver_strategies)*len(weight_constants))
    print('Total time approx: ', (sum(number_of_mu_train)*len(beta_values)+number_of_mu_test+len(mu_validation)+ len(beta_values)*(sum(number_of_mu_train)+len(number_of_mu_train)*(number_of_mu_test+len(mu_validation)))*len(svd_truncation_tolerances)*len(rom_solver_strategies)*len(weight_constants))*5,'sec')
    print('Total time approx: ', (sum(number_of_mu_train)*len(beta_values)+number_of_mu_test+len(mu_validation)+ len(beta_values)*(sum(number_of_mu_train)+len(number_of_mu_train)*(number_of_mu_test+len(mu_validation)))*len(svd_truncation_tolerances)*len(rom_solver_strategies)*len(weight_constants))*5/(60),'mins')
    print('Total time approx: ', (sum(number_of_mu_train)*len(beta_values)+number_of_mu_test+len(mu_validation)+ len(beta_values)*(sum(number_of_mu_train)+len(number_of_mu_train)*(number_of_mu_test+len(mu_validation)))*len(svd_truncation_tolerances)*len(rom_solver_strategies)*len(weight_constants))*5/(60*60),'hrs')
    print('Total time approx: ', (sum(number_of_mu_train)*len(beta_values)+number_of_mu_test+len(mu_validation)+ len(beta_values)*(sum(number_of_mu_train)+len(number_of_mu_train)*(number_of_mu_test+len(mu_validation)))*len(svd_truncation_tolerances)*len(rom_solver_strategies)*len(weight_constants))*5/(60*60*24),'days')
    print(':::::::::::::::::::::::::::::::::::::::::')
    input('Pause')

    if update_parameters:
        KratosMultiphysics.kratos_utilities.DeleteDirectoryIfExisting('Mu_parameters')
        get_multiple_parameters(number_of_train_values = number_of_mu_train, 
                                number_of_test_values  = number_of_mu_test , 
                                mu_validation          = mu_validation     ,
                                angle_range            = angle_range       , 
                                mach_range             = mach_range        , 
                                method                 = 'Halton'          ,
                                alpha_values = alpha_values, beta_values = beta_values)

    mu_test  = load_mu_parameters('mu_test') 
    mu_validation = load_mu_parameters('mu_validation')

    general_rom_manager_parameters = GetRomManagerParameters()
    rom_manager = RomManager(project_parameters_name,general_rom_manager_parameters,
                            CustomizeSimulation,UpdateProjectParameters,UpdateMaterialParametersFile,
                            relaunch_FOM=relaunch_FOM, relaunch_ROM=relaunch_ROM, mu_names=['Alpha','Mach','Weight','Beta','MuId','SVDTolerance','RomSolverStrategy'],
                            rebuild_phi=rebuild_phi)

    for rom_solver_strategy in rom_solver_strategies:

        for id, mu in enumerate(number_of_mu_train):

            for alpha, beta in zip(alpha_values, beta_values):


                for svd_truncation_tolerance in svd_truncation_tolerances:

                    for weight_constant in weight_constants:

                        resume = []; run_error = 0.0; case = 'ROMvsFOM'; modes = 0

                        rom_manager.general_rom_manager_parameters["ROM"]["svd_truncation_tolerance"].SetDouble(svd_truncation_tolerance)
                        
                        mu_train = load_mu_parameters(f'mu_train_{id}_{alpha}_{beta}') 
                        for sublist in mu_train:
                            sublist[2:] = [weight_constant,beta,id,svd_truncation_tolerance,rom_solver_strategy]
                        for sublist in mu_test:
                            sublist[2:] = [weight_constant,beta,id,svd_truncation_tolerance,rom_solver_strategy]
                        for sublist in mu_validation:
                            sublist[2:] = [weight_constant,beta,id,svd_truncation_tolerance,rom_solver_strategy]

                        rom_manager.Fit(mu_train=mu_train, validate=False)

                        KratosMultiphysics.kratos_utilities.DeleteFileIfExisting('selected_elements_list.txt')
                        extract_unique_elements(mu_train)
                        
                        rom_manager.Fit(mu_train=mu_train)
                        
                        if TEST:
                            rom_manager.Test(mu_test)

                        if VALIDATION:
                            rom_manager.RunFOM(mu_run=mu_validation)
                            training_stages = rom_manager.general_rom_manager_parameters["rom_stages_to_test"].GetStringArray()
                            if any(item == "ROM" for item in training_stages):
                                rom_manager.RunROM(mu_run=mu_validation)
                        
                        if os.path.exists('rom_data/RightBasisMatrix.npy'): 
                            modes = np.load('rom_data/RightBasisMatrix.npy').shape[1]
                        
                        if rom_solver_strategy: 
                            strategy = 'Newton Raphson'
                        else:
                            strategy = 'Line Search'

                        resume.append([case, id, svd_truncation_tolerance, modes, weight_constant, beta, strategy,
                                        len(mu_train), rom_manager.ROMvsFOM['Fit'], 
                                        len(mu_test), rom_manager.ROMvsFOM['Test'], 
                                        len(mu_validation), rom_manager.ROMvsFOM['Run']])

                        if os.path.exists('resume.xlsx'):
                            wb = openpyxl.load_workbook('resume.xlsx')
                            hoja = wb.active
                            for item in resume:
                                hoja.append(item)
                            wb.save('resume.xlsx')
                        else:
                            wb = openpyxl.Workbook()
                            hoja = wb.active
                            hoja.append(('Case', 'Mu id', 'SVD tol', 'Modes', 'Weight constant', 
                                        'Density constant', 'Rom solver strategy', 'Fit cases', 'Error', 'Test cases', 
                                        'Error', 'Validation cases', 'Error'))
                            for item in resume:
                                hoja.append(item)
                            wb.save('resume.xlsx')