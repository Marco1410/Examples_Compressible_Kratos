import os
import time
import openpyxl
import numpy as np
import matplotlib
import matplotlib.patheffects as path_effects
from scipy.optimize import minimize_scalar
import matplotlib.pyplot as plt
matplotlib.use('Agg')
from scipy.stats import qmc  
import KratosMultiphysics
import KratosMultiphysics.kratos_utilities
import KratosMultiphysics.CompressiblePotentialFlowApplication as CPFApp
from KratosMultiphysics.RomApplication.rom_manager import RomManager
from KratosMultiphysics.RomApplication.calculate_rom_basis_output_process import CalculateRomBasisOutputProcess


# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
#
# get multiple parameters by Halton or LatinHypercube methods
#
def get_multiple_parameters(regions=[], number_test_values=0,
                            angle=[], mach=[], mu_validation=[], 
                            method='Halton', update_mu_test = True, alpha=1.0, beta=1.0):
    if method == 'Halton':
        sampler = qmc.Halton(d=2)
    elif method == 'LatinHypercube':
        sampler = qmc.LatinHypercube(d=2)
    mu_train = []; mu_test = []
    mu_train_not_scaled = []; mu_test_not_scaled = []; mu_validation_not_scaled = []
    for mach_range, alpha_range, number_of_values in regions:
        mach_common_min  = max(mach_range[0] , mach[0])
        mach_common_max  = min(mach_range[1] , mach[1])
        angle_common_min = max(alpha_range[0], angle[0])
        angle_common_max = min(alpha_range[1], angle[1])
        if mach_common_min < mach_common_max and angle_common_min < angle_common_max and number_of_values > 0:
            sample = sampler.random(number_of_values)
            # Density transformation: Alpha: density X; Beta:  density Y
            transformed_points = np.zeros_like(sample)
            transformed_points[:, 0] = sample[:, 0]**beta
            transformed_points[:, 1] = sample[:, 1]**alpha
            values = qmc.scale(transformed_points, [angle_common_min,mach_common_min], [angle_common_max,mach_common_max])
            for i in range(number_of_values):
                #Angle of attack , Mach infinit
                mu_train.append([values[i,0], values[i,1]])
                mu_train_not_scaled.append([transformed_points[i,0], transformed_points[i,1]])
    np.save('Mu_history/mu_train', mu_train)
    np.save('Mu_history/mu_train_not_scaled', mu_train_not_scaled)
    if number_test_values > 0 and update_mu_test:
        sample = sampler.random(number_test_values)
        values = qmc.scale(sample, [angle_common_min,mach_common_min], [angle_common_max,mach_common_max])
        for i in range(number_test_values):
            #Angle of attack , Mach infinit
            mu_test.append([values[i,0], values[i,1]])
            mu_test_not_scaled.append([sample[i,0], sample[i,1]])
        np.save('Mu_history/mu_test', mu_test)
        np.save('Mu_history/mu_test_not_scaled', mu_test_not_scaled)
    if os.path.exists("Mu_history/mu_test.npy") and not update_mu_test:
        mu_test = np.load('Mu_history/mu_test.npy')
        mu_test_not_scaled = np.load('Mu_history/mu_test_not_scaled.npy')
        mu_test =  [mu.tolist() for mu in mu_test]
        mu_test_not_scaled =  [mu.tolist() for mu in mu_test_not_scaled]
    if len(mu_validation)>0:
        mu_validation_not_scaled = get_not_scale_parameters(mu_validation, angle, mach)
        np.save('Mu_history/mu_validation', mu_validation)
        np.save('Mu_history/mu_validation_not_scaled', mu_validation_not_scaled)
    plot_mu_values(mu_train, mu_test, mu_validation, 'MuValues')    
    return mu_train, mu_test, mu_train_not_scaled, mu_test_not_scaled, mu_validation_not_scaled

def get_not_scale_parameters(mu=[None], angle=[], mach=[]):
    mu_not_scaled = []
    if len(mu) > 0:
        sample = np.array(mu)
        values = qmc.scale(sample, [angle[0],mach[0]], [angle[1],mach[1]], reverse=True)
        for i in range(len(mu)):
            #Angle of attack , Mach infinit
            mu_not_scaled.append([values[i,0], values[i,1]])    
    return mu_not_scaled

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
#
# plot parameters
#
def plot_mu_values(mu_train, mu_test, mu_validation, name):
    fig, ax = plt.subplots()
    fig.set_figwidth(10.0)
    fig.set_figheight(6.0)
    mu_train = np.array(mu_train).reshape(-1, 2) if len(mu_train) > 0 else np.empty((0, 2))
    mu_test = np.array(mu_test).reshape(-1, 2) if len(mu_test) > 0 else np.empty((0, 2))
    mu_validation = np.array(mu_validation).reshape(-1, 2) if len(mu_validation) > 0 else np.empty((0, 2))
    if len(mu_train) + len(mu_test) + len(mu_validation) > 0:
        all_mu_values = np.concatenate((mu_train, mu_test, mu_validation), axis=0)
        min_alpha, min_mach = np.min(all_mu_values[:, 0]), np.min(all_mu_values[:, 1])
        max_alpha, max_mach = np.max(all_mu_values[:, 0]), np.max(all_mu_values[:, 1])
        regions = [
            ((min_mach, min_alpha), (max_mach, max_alpha), 'red')
        ]
        for bottom_left, top_right, color in regions:
            rect = plt.Rectangle(
                bottom_left,
                top_right[0] - bottom_left[0],
                top_right[1] - bottom_left[1],
                facecolor=color, edgecolor=color, alpha=0.25
            )
            ax.add_patch(rect)
    def plot_points_with_labels(data, marker, color, label, dx=0.0, dy=-0.08):
        if len(data) > 0:
            for i, (alpha, mach) in enumerate(data):
                ax.plot(mach, alpha, marker, color=color, label=label if i == 0 else "", markersize=8)
                text = ax.text(mach + dx, alpha + dy, str(i), fontsize=8, ha='center', va='center', color=color)
                text.set_path_effects([
                    path_effects.Stroke(linewidth=1.5, foreground='black'),
                    path_effects.Normal()
                ])
    plot_points_with_labels(mu_train, 's', 'blue', "Train Values")
    plot_points_with_labels(mu_test, 'x', 'red', "Test Values")
    plot_points_with_labels(mu_validation, '*', 'green', "Validation Values")
    plt.title('Mu Values')
    plt.ylabel('Alpha')
    plt.xlabel('Mach')
    plt.grid(True)
    plt.legend(bbox_to_anchor=(1.02, 1), loc='upper left', borderaxespad=0.)
    plt.tight_layout()
    plt.savefig(f"Mu_history/{name}.pdf", bbox_inches='tight', dpi=400)
    plt.close('all')

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
#
# load parameters
#
def load_mu_parameters(name):
    filename = f'Mu_history/{name}.npy'
    if os.path.exists(filename):
        mu_npy = np.load(filename)
        mu =  [mu.tolist() for mu in mu_npy]
    else:
        mu = []
    return mu

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

def CustomizeSimulation(cls, global_model, parameters, mu, constant):

    class CustomSimulation(cls):

        def __init__(self, model,project_parameters):
            super().__init__(model,project_parameters)
            if len(mu) > 0:
                self.case_name = f'{mu[0]}, {mu[1]}'
                self.case_subname = f'{constant[0]}_{constant[1]}' if constant[1] > 0 else f'{constant[0]}'

            if parameters["output_processes"].Has("gid_output"):
                self.simulation_name = parameters["output_processes"]["gid_output"][0]["Parameters"]["output_name"].GetString().removeprefix('Results/')

                if not 'FOM' in self.simulation_name:
                    parameters["output_processes"]["gid_output"][0]["Parameters"]["output_name"].SetString(f'Results/{self.case_subname}/{self.simulation_name}')
                    vtk_output_name = parameters["output_processes"]["vtk_output"][0]["Parameters"]["output_path"].GetString().removeprefix('Results/')
                    parameters["output_processes"]["vtk_output"][0]["Parameters"]["output_path"].SetString(f'Results/{self.case_subname}/{vtk_output_name}')

                # if 'ROM' in self.simulation_name: # Use Neton Raphson for rom simulations
                #     parameters["solver_settings"]["solving_strategy_settings"]["type"].SetString("newton_raphson")
            
        def InitializeSolutionStep(self):
            super().InitializeSolutionStep()

        def FinalizeSolutionStep(self):
            super().FinalizeSolutionStep()

            non_historical_gradient_variable = True
            KratosMultiphysics.ComputeNodalGradientProcess(
                    self.model["MainModelPart"],
                    KratosMultiphysics.DENSITY,
                    KratosMultiphysics.DISTANCE_GRADIENT,
                    KratosMultiphysics.NODAL_AREA,
                    non_historical_gradient_variable).Execute()

        def Run(self):

            start_time = time.time()
            self.Initialize()

            for elem in self.model["MainModelPart"].Elements: #Initilize of elemental variable
                elem.SetValue(KratosMultiphysics.ACTIVATION_LEVEL, 0)
            
            if 'FOM' in self.simulation_name: #MARK ELEMENTS
                self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.REFINEMENT_LEVEL] = 1

            if 'ROM' in self.simulation_name: #USE WEIGHT VERSION
                self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.REFINEMENT_LEVEL] =  1 if constant[1] > 0 else 0

            # SET WEIGHT CONSTANT
            if 'ROM' in self.simulation_name:
                self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.CONSTRAINT_SCALE_FACTOR] = constant[1]

            self.RunSolutionLoop()
            self.Finalize()

            exe_time = time.time() - start_time

            if parameters["output_processes"].Has("gid_output"):
                info_steps_list = []
                u_error = 0 
                modes   = 0

                for process in self._GetListOfOutputProcesses():
                    if isinstance(process, CalculateRomBasisOutputProcess):
                        BasisOutputProcess = process

                if 'FOM' in self.simulation_name:
                    case = 'FOM'
                    if 'Fit' in self.simulation_name:
                        case_type = 'train_fom'
                    elif 'Test' in self.simulation_name:
                        case_type = 'test_fom'
                    elif 'Run' in self.simulation_name:
                        case_type = 'run_fom' 

                    fom = BasisOutputProcess._GetSnapshotsMatrix()

                    np.save(f'{case}_Snapshots/{self.case_name}',fom)

                    if case_type == 'train_fom': #Only for train FOM
                        modelpart = self.model["MainModelPart"]
                        fout = open("selected_elements_list.txt",'a') #Elements list for rom approximation
                        for elem in modelpart.Elements:
                            if elem.GetValue(KratosMultiphysics.ACTIVATION_LEVEL):
                                fout.write("%s\n" %(elem.Id))
                        fout.close()
                    
                elif 'ROM' in self.simulation_name:
                    case = 'ROM'
                    if 'Fit' in self.simulation_name:
                        case_type = 'train_rom'
                    elif 'Test' in self.simulation_name:
                        case_type = 'test_rom' 
                    elif 'Run' in self.simulation_name:
                        case_type = 'run_rom' 

                    u_rom = BasisOutputProcess._GetSnapshotsMatrix()

                    u_fom = np.load(f'FOM_Snapshots/{self.case_name}.npy')

                    if not os.path.exists(f'{case}_Snapshots/{self.case_subname}'):
                        os.mkdir(f'{case}_Snapshots/{self.case_subname}')
                    np.save(f'{case}_Snapshots/{self.case_subname}/{self.case_name}', u_rom)
                                        
                    u_error = np.linalg.norm(u_fom-u_rom)/np.linalg.norm(u_fom)

                if os.path.exists('rom_data/RightBasisMatrix.npy'): 
                    modes = np.load('rom_data/RightBasisMatrix.npy').shape[1]

                if not 'FOM' in self.simulation_name:
                    if not os.path.exists(f"{case}_Skin_Data/{self.case_subname}"):
                        os.mkdir(f"{case}_Skin_Data/{self.case_subname}")

                    skin_data_filename = f"{case}_Skin_Data/{self.case_subname}/{self.case_name}.dat"
                else:
                    skin_data_filename = f"{case}_Skin_Data/{self.case_name}.dat"

                fout = open(skin_data_filename,'w')
                modelpart = self.model["MainModelPart.Body2D_Body"]
                for node in modelpart.Nodes:
                    x = node.X ; y = node.Y ; z = node.Z
                    cp = node.GetValue(KratosMultiphysics.PRESSURE_COEFFICIENT)
                    fout.write("%s %s %s %s\n" %(x,y,z,cp))
                fout.close()

                info_steps_list.append([constant[0],
                                        constant[1],
                                        case_type,
                                        mu[0],
                                        mu[1], 
                                        self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.RESIDUAL_NORM],
                                        self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.NL_ITERATION_NUMBER],
                                        u_error,
                                        modes,
                                        round(exe_time, 2),
                                        self.model["MainModelPart"].NumberOfNodes(),
                                        self.model["MainModelPart"].ProcessInfo[CPFApp.LIFT_COEFFICIENT],
                                        self.model["MainModelPart"].ProcessInfo[KratosMultiphysics.DRAG_COEFFICIENT]])
                
                if os.path.exists(f'{case}_data.xlsx'):
                    wb = openpyxl.load_workbook(f'{case}_data.xlsx')
                    hoja = wb.active
                    for item in info_steps_list:
                        hoja.append(item)
                    wb.save(f'{case}_data.xlsx')
                else:
                    wb = openpyxl.Workbook()
                    hoja = wb.active
                    hoja.append(('Strategy', 'Constant', 'Case name','Angle [º]', 'Mach', 'Residual norm', 'NL iter','u relative error', 'Modes', 'Time [sec]', 'Nodes', 'Cl', 'Cd'))
                    for item in info_steps_list:
                        hoja.append(item)
                    wb.save(f'{case}_data.xlsx')

    return CustomSimulation(global_model, parameters)


# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

def UpdateProjectParameters(parameters, mu=None):
    angle_of_attack = mu[0]
    mach_infinity   = mu[1]
    parameters["processes"]["boundary_conditions_process_list"][0]["Parameters"]["angle_of_attack"].SetDouble(np.double(angle_of_attack))
    parameters["processes"]["boundary_conditions_process_list"][0]["Parameters"]["mach_infinity"].SetDouble(np.double(mach_infinity))

    return parameters

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

def UpdateMaterialParametersFile(material_parametrs_file_name, mu):
    pass

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

def GetRomManagerParameters():
    general_rom_manager_parameters = KratosMultiphysics.Parameters("""{
            "rom_stages_to_train" : ["ROM"],            // ["ROM","HROM"]
            "rom_stages_to_test"  : ["ROM"],            // ["ROM","HROM"]
            "paralellism" : null,                       // null, TODO: add "compss"
            "projection_strategy": "galerkin",          // "lspg", "galerkin", "petrov_galerkin"
            "type_of_decoder" : "linear",               // "linear" "ann_enhanced",  TODO: add "quadratic"
            "assembling_strategy": "global",            // "global", "elemental"
            "save_gid_output": true,                    // false, true #if true, it must exits previously in the ProjectParameters.json
            "save_vtk_output": true,                   // false, true #if true, it must exits previously in the ProjectParameters.json
            "output_name": "id",                        // "id" , "mu"
            "store_nonconverged_fom_solutions": true,
            "ROM":{
                "svd_truncation_tolerance": 1e-12,
                "print_singular_values": true,
                "use_non_converged_sols" : false,
                "model_part_name": "MainModelPart",                         // This changes depending on the simulation: Structure, FluidModelPart, ThermalPart #TODO: Idenfity it automatically
                "nodal_unknowns": ["VELOCITY_POTENTIAL", "AUXILIARY_VELOCITY_POTENTIAL"],     // Main unknowns. Snapshots are taken from these
                "rom_basis_output_format": "numpy",
                "rom_basis_output_name": "RomParameters",
                "rom_basis_output_folder": "rom_data",
                "snapshots_control_type": "step",                          // "step", "time"
                "snapshots_interval": 1,
                "galerkin_rom_bns_settings": {
                    "monotonicity_preserving": false
                },
                "lspg_rom_bns_settings": {
                    "train_petrov_galerkin": false,
                    "basis_strategy": "reactions",                        // 'residuals', 'jacobian', 'reactions'
                    "include_phi": false,
                    "svd_truncation_tolerance": 0,
                    "solving_technique": "normal_equations",              // 'normal_equations', 'qr_decomposition'
                    "monotonicity_preserving": false
                },
                "petrov_galerkin_rom_bns_settings": {
                    "monotonicity_preserving": false
                }
            },
            "HROM":{
                "element_selection_type": "empirical_cubature",
                "element_selection_svd_truncation_tolerance": 1e-12,
                "create_hrom_visualization_model_part" : true,
                "echo_level" : 1,                                       
                "hrom_format": "numpy",
                "initial_candidate_elements_model_part_list" : [],
                "initial_candidate_conditions_model_part_list" : [],
                "include_nodal_neighbouring_elements_model_parts_list":[],
                "include_minimum_condition": false,
                "include_condition_parents": true
            }
        }""")

    return general_rom_manager_parameters

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

def evaluate_constant_lspg(constant):
    strategy = 'lspg'
    case = 'ROMvsFOM'
    general_rom_manager_parameters = GetRomManagerParameters()
    project_parameters_name = "ProjectParameters.json"
    general_rom_manager_parameters["projection_strategy"].SetString(strategy)

    rom_manager = RomManager(project_parameters_name,general_rom_manager_parameters,
                            CustomizeSimulation,UpdateProjectParameters,UpdateMaterialParametersFile,
                            relaunch_FOM=False, relaunch_ROM=True, 
                            rebuild_phi=True, constant=[strategy,constant])
    
    rom_manager.Fit(mu_train=load_mu_parameters(f'0_{strategy}_mu_train'))
    rom_manager.Test(load_mu_parameters(f'0_{strategy}_mu_test'))
    
    resume = []
    resume.append([strategy, case, constant, 
                   len(load_mu_parameters(f'0_{strategy}_mu_train')), rom_manager.ROMvsFOM['Fit'],
                   len(load_mu_parameters(f'0_{strategy}_mu_test')), rom_manager.ROMvsFOM['Test']])

    if os.path.exists('resume.xlsx'):
        wb = openpyxl.load_workbook('resume.xlsx')
        hoja = wb.active
        hoja.append(('Evaluation case', ' '))
        hoja.append(('Strategy', 'Case', 'Constant', 'Fit cases', 'Error', 'Test cases', 'Error'))
        for item in resume:
            hoja.append(item)
        wb.save('resume.xlsx')
    else:
        wb = openpyxl.Workbook()
        hoja = wb.active
        hoja.append(('Evaluation case', ' '))
        hoja.append(('Strategy', 'Case', 'Constant', 'Fit cases', 'Error', 'Test cases', 'Error'))
        for item in resume:
            hoja.append(item)
        wb.save('resume.xlsx')
        
    if rom_manager.ROMvsFOM['Fit'] < 1e-9:
        return rom_manager.ROMvsFOM['Test']
    else:
        return rom_manager.ROMvsFOM['Fit']

def evaluate_constant_galerkin(constant):
    strategy = 'galerkin'
    case = 'ROMvsFOM'
    general_rom_manager_parameters = GetRomManagerParameters()
    project_parameters_name = "ProjectParameters.json"
    general_rom_manager_parameters["projection_strategy"].SetString(strategy)

    rom_manager = RomManager(project_parameters_name,general_rom_manager_parameters,
                            CustomizeSimulation,UpdateProjectParameters,UpdateMaterialParametersFile,
                            relaunch_FOM=False, relaunch_ROM=True, 
                            rebuild_phi=True, constant=[strategy,constant])

    rom_manager.Fit(mu_train=load_mu_parameters(f'0_{strategy}_mu_train'))
    rom_manager.Test(load_mu_parameters(f'0_{strategy}_mu_test'))

    resume = []
    resume.append([strategy, case, constant, 
                   len(load_mu_parameters(f'0_{strategy}_mu_train')), rom_manager.ROMvsFOM['Fit'],
                   len(load_mu_parameters(f'0_{strategy}_mu_test')), rom_manager.ROMvsFOM['Test']])

    if os.path.exists('resume.xlsx'):
        wb = openpyxl.load_workbook('resume.xlsx')
        hoja = wb.active
        hoja.append(('Evaluation case', ' '))
        hoja.append(('Strategy', 'Case', 'Constant', 'Fit cases', 'Error', 'Test cases', 'Error'))
        for item in resume:
            hoja.append(item)
        wb.save('resume.xlsx')
    else:
        wb = openpyxl.Workbook()
        hoja = wb.active
        hoja.append(('Evaluation case', ' '))
        hoja.append(('Strategy', 'Case', 'Constant', 'Fit cases', 'Error', 'Test cases', 'Error'))
        for item in resume:
            hoja.append(item)
        wb.save('resume.xlsx')
        
    if rom_manager.ROMvsFOM['Fit'] < 1e-9:
        return rom_manager.ROMvsFOM['Test']
    else:
        return rom_manager.ROMvsFOM['Fit']

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #



if __name__ == "__main__":

    folder_names = ["FOM_Snapshots"  , "FOM_Skin_Data", 
                    "ROM_Snapshots"  , "ROM_Skin_Data", 
                    "RBF_Snapshots"  , "RBF_Skin_Data", 
                    "Train_Captures" , "Test_Captures", "Validation", "Mu_history"]
    for name in folder_names:
        if not os.path.exists(name):
            os.mkdir(name)

    ################################
    # PARAMETERS SETTINGS
    clean_data         = False
    update_parameters  = True
    update_mu_test     = True
    VALIDATION         = True
    number_of_mu_train = 3
    number_of_mu_test  = 1
    alpha              = 0.75
    beta               = 0.75
    strategies         = ['galerkin']
    # ADAPTIVE SAMPLING #############
    adapt              = True
    initial_fit        = True
    tolerances_error   = [0.8,0.5,0.2,0.1,0.01]
    # OPTIMIZATION ##################
    optimize           = False
    optimal_constants  = [-1]
    constant_range     = [1e-6, 10]
    tolerance          = 1e-9
    iterations         = 15
    mu_prefix          = f'0_{strategies[0]}_'
    #################################
    mach_range         = [0.70, 0.73]
    angle_range        = [0.00, 2.00]
    relaunch_FOM       = False 
    relaunch_ROM       = True
    rebuild_phi        = True 
    ################################

    regions = [
        ((mach_range),(angle_range), number_of_mu_train)  
    ]

    mu_validation = []
    mu_validation_not_scaled = []
    if VALIDATION:
        if angle_range[0] <= 0.2 and angle_range[1]>= 0.2 and mach_range[0] <= 0.75 and mach_range[1]>=0.75:
            mu_validation.append([0.2,0.75])
        if angle_range[0] <= 0.7 and angle_range[1]>= 0.7 and mach_range[0] <= 0.75 and mach_range[1]>=0.75:
            mu_validation.append([0.7,0.75])
        if angle_range[0] <= 1.0 and angle_range[1]>= 1.0 and mach_range[0] <= 0.70 and mach_range[1]>=0.70:
            mu_validation.append([1.0,0.70])
        if angle_range[0] <= 2.0 and angle_range[1]>= 2.0 and mach_range[0] <= 0.70 and mach_range[1]>=0.70:
            mu_validation.append([2.0,0.70])
    
    if update_parameters:
        (mu_train, mu_test,
        mu_train_not_scaled, mu_test_not_scaled,
        mu_validation_not_scaled) = get_multiple_parameters(regions             = regions           ,
                                                            number_test_values  = number_of_mu_test , 
                                                            angle               = angle_range       , 
                                                            mach                = mach_range        , 
                                                            mu_validation       = mu_validation     ,
                                                            method              = 'Halton'          ,
                                                            update_mu_test      = update_mu_test    ,
                                                            alpha = alpha, beta = beta )
        KratosMultiphysics.kratos_utilities.DeleteFileIfExisting('selected_elements_list.txt')
        KratosMultiphysics.kratos_utilities.DeleteFileIfExisting('ROM_data.xlsx')
        KratosMultiphysics.kratos_utilities.DeleteFileIfExisting('resume.xlsx')
    else:
        if clean_data:
            KratosMultiphysics.kratos_utilities.DeleteFileIfExisting('ROM_data.xlsx')
            KratosMultiphysics.kratos_utilities.DeleteFileIfExisting('resume.xlsx')
            KratosMultiphysics.kratos_utilities.DeleteDirectoryIfExisting('Train_Captures')
            KratosMultiphysics.kratos_utilities.DeleteDirectoryIfExisting('Test_Captures')
            KratosMultiphysics.kratos_utilities.DeleteDirectoryIfExisting('Validation')
        for name in folder_names:
            if not os.path.exists(name):
                os.mkdir(name)
        if mu_prefix == '':
            mu_train = load_mu_parameters('mu_train')
            mu_train_not_scaled = load_mu_parameters('mu_train_not_scaled')
            mu_test = load_mu_parameters('mu_test')
            mu_test_not_scaled = load_mu_parameters('mu_test_not_scaled')
        else:
            mu_train = load_mu_parameters(f'{mu_prefix}mu_train')
            mu_train_not_scaled = load_mu_parameters(f'{mu_prefix}mu_train_not_scaled')
            mu_test = load_mu_parameters(f'{mu_prefix}mu_test')
            mu_test_not_scaled = load_mu_parameters(f'{mu_prefix}mu_test_not_scaled')
        mu_validation = load_mu_parameters('mu_validation')
        mu_validation_not_scaled = load_mu_parameters('mu_validation_not_scaled')
        plot_mu_values(mu_train, mu_test, mu_validation, 'MuValues')

    if adapt:
        for strategy in strategies:

            resume = []; run_error = 0.0; case = 'ROMvsFOM'

            general_rom_manager_parameters = GetRomManagerParameters()
            project_parameters_name = "ProjectParameters.json"
            general_rom_manager_parameters["projection_strategy"].SetString(strategy)

            rom_manager = RomManager(project_parameters_name,general_rom_manager_parameters,
                                    CustomizeSimulation,UpdateProjectParameters,UpdateMaterialParametersFile,
                                    relaunch_FOM=relaunch_FOM, relaunch_ROM=relaunch_ROM, 
                                    rebuild_phi=rebuild_phi, constant=[strategy, 0])
            
            for i_tol, tolerance in enumerate(tolerances_error):
                if mu_test==[]: break
                while mu_test:
                    if initial_fit:
                        rom_manager.Fit(validate=False, mu_train=mu_train)                            
                        rom_manager.Test(mu_test)

                    fom_snapshots = rom_manager.data_base.get_snapshots_matrix_from_database(mu_test, table_name=f'FOM')
                    rom_snapshots = rom_manager.data_base.get_snapshots_matrix_from_database(mu_test, table_name=f'ROM')
                    errors = np.linalg.norm(fom_snapshots - rom_snapshots, axis=0) / np.linalg.norm(fom_snapshots, axis=0)

                    new_mu_train = []; new_mu_train_not_scaled = []
                    for j, error in enumerate(errors):
                        if error > tolerance:
                            new_mu_train.append(mu_test[j])
                            new_mu_train_not_scaled.append(mu_test_not_scaled[j])

                    if len(new_mu_train)>0:
                        # Update mu_train mu_test
                        mu_train.extend(new_mu_train)
                        mu_train_not_scaled.extend(new_mu_train_not_scaled)
                        mu_test = [mu for i, mu in enumerate(mu_test) if errors[i] <= tolerance]
                        mu_test_not_scaled = [mu for i, mu in enumerate(mu_test_not_scaled) if errors[i] <= tolerance]
                        np.save(f'Mu_history/{i_tol}_{strategy}_mu_train', mu_train)
                        np.save(f'Mu_history/{i_tol}_{strategy}_mu_train_not_scaled', mu_train_not_scaled)
                        np.save(f'Mu_history/{i_tol}_{strategy}_mu_test', mu_test)        
                        np.save(f'Mu_history/{i_tol}_{strategy}_mu_test_not_scaled', mu_test_not_scaled)        
                        plot_mu_values(mu_train, mu_test, mu_validation, f'{i_tol}_{strategy}_MuValues')
                        initial_fit = True
                    else:
                        initial_fit = False
                        break

            rom_manager.Fit(mu_train=mu_train)
            rom_manager.Test(mu_test=mu_test)

            if VALIDATION:
                rom_manager.RunFOM(mu_run=mu_validation)
                training_stages = rom_manager.general_rom_manager_parameters["rom_stages_to_test"].GetStringArray()
                if any(item == "ROM" for item in training_stages):
                    rom_manager.RunROM(mu_run=mu_validation)

                for mu in mu_validation:
                    run_fom = []
                    run_rom = []
                    if os.path.exists(f'FOM_Snapshots/{mu[0]}, {mu[1]}.npy'):
                        run_fom.append(np.load(f'FOM_Snapshots/{mu[0]}, {mu[1]}.npy'))
                    if os.path.exists(f'ROM_Snapshots/{strategy}/{mu[0]}, {mu[1]}.npy'):
                        run_rom.append(np.load(f'ROM_Snapshots/{strategy}/{mu[0]}, {mu[1]}.npy'))
                
                run_fom = np.block(run_fom)
                if any(item == "ROM" for item in training_stages):
                    run_rom = np.block(run_rom)

            if VALIDATION:
                if any(item == "ROM" for item in training_stages):
                    run_error = np.linalg.norm(run_fom-run_rom)/np.linalg.norm(run_fom)
            
            resume.append([strategy, case, len(mu_train), rom_manager.ROMvsFOM['Fit'], len(mu_test), rom_manager.ROMvsFOM['Test'], len(mu_validation), run_error])

            if os.path.exists('resume.xlsx'):
                wb = openpyxl.load_workbook('resume.xlsx')
                hoja = wb.active
                hoja.append(('Adaptive sampling','Run case'))
                hoja.append(('Strategy', 'Case', 'Fit cases', 'Error', 'Test cases', 'Error', 'Validation cases', 'Error'))
                for item in resume:
                    hoja.append(item)
                wb.save('resume.xlsx')
            else:
                wb = openpyxl.Workbook()
                hoja = wb.active
                hoja.append(('Adaptive sampling','Run case'))
                hoja.append(('Strategy', 'Case', 'Fit cases', 'Error', 'Test cases', 'Error', 'Validation cases', 'Error'))
                for item in resume:
                    hoja.append(item)
                wb.save('resume.xlsx')

    if optimize:
        for strategy, optimal_constant in zip(strategies, optimal_constants):
            if optimal_constant < 0:
                # Error minimization
                if strategy=='galerkin':
                    if not os.path.exists(f'Results/galerkin_1.0'):
                        evaluate_constant_galerkin(constant=1.0)
                    result = minimize_scalar(evaluate_constant_galerkin, bounds=constant_range, method='bounded', tol=tolerance, options={'maxiter':iterations})
                else:
                    if not os.path.exists(f'Results/lspg_1.0'):
                        evaluate_constant_lspg(constant=1.0)
                    result = minimize_scalar(evaluate_constant_lspg, bounds=constant_range, method='bounded', tol=tolerance, options={'maxiter':iterations})
                optimal_constant = result.x

            resume = []; run_error = 0.0; case = 'ROMvsFOM'

            general_rom_manager_parameters = GetRomManagerParameters()
            project_parameters_name = "ProjectParameters.json"
            general_rom_manager_parameters["projection_strategy"].SetString(strategy)

            rom_manager = RomManager(project_parameters_name,general_rom_manager_parameters,
                                    CustomizeSimulation,UpdateProjectParameters,UpdateMaterialParametersFile,
                                    relaunch_FOM=relaunch_FOM, relaunch_ROM=relaunch_ROM, 
                                    rebuild_phi=rebuild_phi, constant=[strategy,optimal_constant])

            rom_manager.Fit(mu_train=mu_train)
            rom_manager.Test(mu_test)

            if VALIDATION:
                rom_manager.RunFOM(mu_run=mu_validation)
                training_stages = rom_manager.general_rom_manager_parameters["rom_stages_to_test"].GetStringArray()
                if any(item == "ROM" for item in training_stages):
                    rom_manager.RunROM(mu_run=mu_validation)

                for mu in mu_validation:
                    run_fom = []
                    run_rom = []
                    if os.path.exists(f'FOM_Snapshots/{mu[0]}, {mu[1]}.npy'):
                        run_fom.append(np.load(f'FOM_Snapshots/{mu[0]}, {mu[1]}.npy'))
                    if os.path.exists(f'ROM_Snapshots/{strategy}_{optimal_constant}/{mu[0]}, {mu[1]}.npy'):
                        run_rom.append(np.load(f'ROM_Snapshots/{strategy}_{optimal_constant}/{mu[0]}, {mu[1]}.npy'))
                
                run_fom = np.block(run_fom)
                if any(item == "ROM" for item in training_stages):
                    run_rom = np.block(run_rom)
                    run_error = np.linalg.norm(run_fom-run_rom)/np.linalg.norm(run_fom)
            
            resume.append([strategy, case, optimal_constant, len(mu_train), rom_manager.ROMvsFOM['Fit'], len(mu_test), rom_manager.ROMvsFOM['Test'], len(mu_validation), run_error])

            if os.path.exists('resume.xlsx'):
                wb = openpyxl.load_workbook('resume.xlsx')
                hoja = wb.active
                hoja.append(('Optimal Constant', 'Run case'))
                hoja.append(('Strategy', 'Case', 'Constant', 'Fit cases', 'Error', 'Test cases', 'Error', 'Validation cases', 'Error'))
                for item in resume:
                    hoja.append(item)
                wb.save('resume.xlsx')
            else:
                wb = openpyxl.Workbook()
                hoja = wb.active
                hoja.append(('Optimal Constant', 'Run case'))
                hoja.append(('Strategy', 'Case', 'Constant', 'Fit cases', 'Error', 'Test cases', 'Error', 'Validation cases', 'Error'))
                for item in resume:
                    hoja.append(item)
                wb.save('resume.xlsx')